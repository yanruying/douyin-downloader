#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
核心功能 - Excel 导出
"""
import os
from datetime import datetime

from douyin_downloader.constants import (
    OPENPYXL_AVAILABLE, Alignment, get_column_letter
)
from douyin_downloader.utils.file_utils import (
    sanitize_filename, safe_mkdir
)

if OPENPYXL_AVAILABLE:
    import openpyxl
else:
    openpyxl = None


def generate_excel_file(all_awemes, nickname, base_folder):
    """
    执行Excel导出（无GUI依赖）。
    成功则返回 filepath，失败则引发 Exception。
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError('[错误] 未安装openpyxl库，请运行: pip install openpyxl')

    try:
        excel_folder = base_folder
        if not safe_mkdir(excel_folder):
            raise OSError('[错误] 无法创建Excel文件夹')

        unique_id = ''
        if all_awemes and len(all_awemes) > 0:
            author_info = all_awemes[0].get('author', {})
            unique_id = author_info.get('unique_id', '') or ''
        
        if unique_id:
            filename = f"{sanitize_filename(nickname)}-{unique_id}.xlsx"
        else:
            filename = f"{sanitize_filename(nickname)}.xlsx"
        filepath = os.path.join(excel_folder, filename)

        wb = openpyxl.Workbook() if openpyxl else None
        if not wb:
            raise Exception('[错误] 无法创建工作簿')
            
        ws = wb.active
        if ws is None:
             raise Exception('[错误] 无法创建工作表')

        ws.title = "作品数据"

        headers = ['类型', '发布时间', '文案', '合集', '点赞数', '评论数', '收藏数', '分享数', '推荐次数', '视频时长', '作品链接']
        ws.append(headers)

        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        for aweme in all_awemes:
            statistics = aweme.get('statistics', {})

            aweme_type = '视频'
            if aweme.get('images'):
                aweme_type = '图集'
            
            create_time = aweme.get('create_time', 0)
            if create_time:
                try:
                    publish_time = datetime.fromtimestamp(create_time).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    publish_time = ''
            else:
                publish_time = ''
            
            mix_name = ''
            mix_info = aweme.get('mix_info', {})
            if isinstance(mix_info, dict):
                mix_name = mix_info.get('mix_name', '') or mix_info.get('mix_name_str', '') or ''
            if not mix_name:
                mix_name = aweme.get('mix_name', '') or aweme.get('mix_name_str', '') or ''
            
            duration_text = ''
            video_info = aweme.get('video', {})
            if isinstance(video_info, dict):
                duration = video_info.get('duration', 0)
                if duration > 0:
                    total_seconds = duration // 1000
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    
                    if hours > 0:
                        duration_text = f"{hours}小时{minutes}分钟{seconds}秒"
                    elif minutes > 0:
                        duration_text = f"{minutes}分钟{seconds}秒"
                    else:
                        duration_text = f"{seconds}秒"
            
            aweme_id = aweme.get('aweme_id', '')
            if aweme_id:
                if aweme.get('images'):
                    link = f"https://www.douyin.com/note/{aweme_id}"
                else:
                    link = f"https://www.douyin.com/video/{aweme_id}"
            else:
                link = ''
            
            row_data = [
                aweme_type,  # 类型
                publish_time,  # 发布时间
                aweme.get('desc', ''),  # 文案
                mix_name,  # 合集
                statistics.get('digg_count', 0),  # 点赞数
                statistics.get('comment_count', 0),  # 评论数
                statistics.get('collect_count', 0),  # 收藏数
                statistics.get('share_count', 0),  # 分享数
                statistics.get('recommend_count', 0),  # 推荐次数
                duration_text,  # 视频时长
                link  # 作品链接
            ]
            ws.append(row_data)
            # 写入后立即设置该行样式，避免二次遍历所有单元格
            row_idx = ws.max_row
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = alignment

        # 自动调整列宽
        for idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(idx)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            if idx == 1:  # 类型列
                adjusted_width = 5
            elif idx == 4:  # 合集列
                adjusted_width = 12
            elif idx == 9:  # 推荐次数列
                adjusted_width = 9
            elif idx == 10:  # 视频时长列
                adjusted_width = 12
            elif idx == 3:  # 文案列
                adjusted_width = 80
            else:
                adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)
        return filepath

    except Exception as e:
        raise Exception(f'[错误] 导出Excel失败: {str(e)}')