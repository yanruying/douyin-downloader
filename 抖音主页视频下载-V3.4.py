#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
抖音主页作品批量下载工具 V3.4
作者：颜如嘤-YanRuYing

功能：
1.批量下载抖音用户主页的所有作品(视频、图片、实况图)
2.导出用户作品信息到Excel表格

本项目完全开源
仅用于学习与研究
禁止将本程序用于任何商业或违法用途
"""

# ========================================
# 模块导入
# ========================================
import os
import re
import sys
import time
import json
import hashlib
import tempfile
import threading
from datetime import datetime

import requests
import configparser
from urllib.parse import unquote, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import partial
try:
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # 定义占位符以便后续代码引用
    Alignment = None
    get_column_letter = None


# ========================================
# 常量定义 - 界面文本
# ========================================
TEXT_APP_NAME = "抖音主页作品批量下载 V3.4 - YanRuYing"

TEXT_INFO_USER = "[信息] 抖音用户：{nickname}"
TEXT_INFO_FETCH = "[进程] 正在获取作品数据..."
TEXT_INFO_FETCH_PAGE = "[进程] 收到第 {page} 页，作品数: {total}【+ {count}】"
TEXT_INFO_FETCH_DONE = "[完成] 获取完成"
TEXT_INFO_DOWNLOAD_START = "[下载] 开始下载 (视频 {vcount}, 图片+实况 {icount}, 线程数 {threads})"

TEXT_WARN_PROFILE_FAIL = "[错误] 获取用户信息异常: Cookie 错误，请重新获取"
TEXT_WARN_FETCH_FAIL = "[错误] 获取第一页失败: {error}"
TEXT_WARN_PAGE_FAIL = "[警告] 第 {page} 页请求异常: {error}"

MAX_DESC_LENGTH = 60

ICON_BYTES = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x10\x00\x00\x00\x10\x08\x06\x00\x00\x00\x1f\xf3\xffa\x00\x00\x00\x01sRGB\x00\xae\xce\x1c\xe9\x00\x00\x01\x15IDAT8Ocd\x80\x80\xffP\x9aT\x8a\x91\x91\x02\xcd`\xcb\x08\x1bp\xfd\x1c\xc4U\x9aFX]G\xd0\x80\xccg\x0f\x19\xa6I\xca1\x88\xb2p0\xbc\xf9\xfb\x13\xc3\x10\xa2\r\x00;\x17\xeccT@\x92\x01 \xadO5|\x19dnn\x81\x9b\x82\xdd\x80\x15\x8b\x18\x18\x82\xfc\x18\x18\xd8\x04\x18`^8h\x1a\xc8`\x7fz=\xc3\x7f\xdd0\x06\xa6+\xab\xf1\x18\xf0\xeb\x03\xc3\x7fV~\xb0\x02\x056n\x06\xaf\x87\xd7\xc1a\x80\xcd\xf9\x98\xb1\xf0\xf3=\xc3\x7f6\x01\x86\xffm\xf3\x18\xfe/\xdf\xc1\xa0ts+\x89\x06@m\x7f\xac\xe1\xc3 ws+\xd8\x150/\x10\xe5\x02\x96\xefo\x19~s\x081\x94\x89j2t\xbf\xb9\x016`\xe7\x97\x8f\x0cn\xdc|Dz\x81\x87\x87\xe1\xff\xe7\xcf`\x8d\xf2\xbf?30\xff\xfd\xcbp\x8fC\x00#\xe0\x90#\x12#\x16\xc4\xc5%\x18^\xbcx\x0eW\xf3\xdf0\x92A\xe4\xf2:\x86w\x7f\x7f\x91\x96\x12\x8d8\x05\x19\xfe\xfd\xff\xcfp\xe1\xc7\x07\xbc\x19\x8c`B"\x94=ai\x93\xec\xec\x0c\x00gfj\x03\xfb\x1e\xc6.\x00\x00\x00\x00IEND\xaeB`\x82'

# ========================================
# 常量定义 - 系统配置
# ========================================
USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36'
REQUEST_TIMEOUT = 12
PAGE_COUNT_PER_REQUEST = 50
DELAY_BETWEEN_PAGES = 0.1

CONFIG_FILE = 'config.ini'

DEFAULT_THREAD_COUNT = 8

# ========================================
# 工具函数 - 文件名和路径处理
# ========================================

def sanitize_filename(filename, max_length=100):
    """清理文件名，移除非法字符"""
    if not filename:
        filename = "unknown"
    # URL解码
    filename = unquote(str(filename))
    # 移除非法字符
    filename = re.sub(r'[\\/*?:"<>|#]', "_", filename)
    # 替换连续空白符
    filename = re.sub(r'\s+', ' ', filename).strip()
    # 截断过长文件名
    if len(filename) > max_length:
        prefix = filename[:max_length // 2 - 2]
        suffix = filename[-(max_length // 2 - 1):]
        filename = f"{prefix}...{suffix}"
    return filename


def safe_mkdir(path):
    """安全创建目录（支持多级目录）"""
    try:
        os.makedirs(path, exist_ok=True)
        return True
    except Exception as e:
        print(f"[错误] 创建目录失败: {path} -> {e}")
        return False


def get_extension_from_url(url, default_ext='.mp4'):
    """从URL提取文件扩展名"""
    try:
        parsed = urlparse(url)
        root, ext = os.path.splitext(parsed.path)
        # 确保扩展名有效
        if ext and 1 < len(ext) <= 6:
            return ext
    except Exception:
        pass
    return default_ext


def generate_unique_filename(base, ext, folder, url):
    """
    生成唯一的文件路径，避免覆盖。
    1. 尝试使用 "描述.ext"
    2. 如果路径过长或已存在，使用 "描述_时间戳_hash.ext"
    3. 如果仍然存在，使用 "描述_时间戳_hash_1.ext", "_2", ...
    4. 如果超过200次，使用 "file_时间戳_hash.ext"
    """
    base_clean = sanitize_filename(base, max_length=150)
    filename = base_clean + ext
    path = os.path.join(folder, filename)
    
    # 检查路径长度是否超限（Windows ~260）或文件是否已存在
    # 预留一些空间，240作为阈值
    if len(path) > 240 or os.path.exists(path):
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        h = hashlib.md5(url.encode('utf-8')).hexdigest()[:8]
        # 截断基础名称以保证总长度
        filename = f"{base_clean[:80]}_{ts}_{h}{ext}"
        path = os.path.join(folder, filename)
    
    # 如果（极罕见情况下）hash 组合后的文件名仍然存在
    counter = 1
    original_path_prefix = path[:-len(ext)]
    while os.path.exists(path):
        filename = f"{original_path_prefix}_{counter}{ext}"
        path = os.path.join(folder, filename)
        counter += 1
        # 极端情况，防止无限循环
        if counter > 200:
            h = hashlib.md5(url.encode('utf-8')).hexdigest()[:8]
            filename = f"file_{int(time.time())}_{h}{ext}"
            path = os.path.join(folder, filename)
            break
    
    return path


def build_expected_filename(desc, ext, is_image, mix_name=None, date_str='', include_date=True):
    """
    构建预期的文件相对路径（用于去重检查）。
    此函数模拟 `download_single_file` 中的文件夹和文件名生成逻辑（但不处理hash或重名）。
    """
    # 根据配置动态构建基础文件名
    base_filename = desc
    if include_date and date_str:
        base_filename = f"{date_str}_{desc}"
    
    filename = sanitize_filename(base_filename, 150) + ext
    folder = ''
    
    if mix_name:
        mix_clean = sanitize_filename(mix_name, max_length=100)
        if is_image:
            folder = os.path.join(mix_clean, 'images')
        else:
            folder = mix_clean
    else:
        if is_image:
            folder = 'images'
        else:
            folder = ''
    
    return os.path.join(folder, filename) if folder else filename


# ========================================
# 配置管理 - INI配置文件读写
# ========================================

def load_config():
    """加载应用配置"""
    cfg = {}
    
    # 1. 尝试从旧版 config.txt (json) 迁移
    old_json = 'config.txt'
    if os.path.exists(old_json):
        try:
            with open(old_json, 'r', encoding='utf-8') as f:
                j = json.load(f) or {}
            if isinstance(j, dict):
                cfg.update(j)
        except Exception:
            pass  # 静默失败

    # 2. 从 config.ini 加载/覆盖
    if os.path.exists(CONFIG_FILE):
        try:
            cp = configparser.ConfigParser(interpolation=None)
            cp.read(CONFIG_FILE, encoding='utf-8')
            
            if 'main' in cp:
                cfg['path'] = cp['main'].get('path', cfg.get('path', ''))
                cfg['cookie'] = cp['main'].get('cookie', cfg.get('cookie', ''))
                
                # 安全地读取布尔值
                try:
                    cfg['use_mix_folder'] = cp['main'].getboolean(
                        'use_mix_folder', 
                        fallback=cfg.get('use_mix_folder', True)
                    )
                except Exception:
                    cfg['use_mix_folder'] = cfg.get('use_mix_folder', True)
                
                try:
                    cfg['include_date_in_filename'] = cp['main'].getboolean(
                        'include_date_in_filename',
                        fallback=cfg.get('include_date_in_filename', True)
                    )
                except Exception:
                    cfg['include_date_in_filename'] = cfg.get('include_date_in_filename', True)
                
                try:
                    cfg['auto_select_after_fetch'] = cp['main'].getboolean(
                        'auto_select_after_fetch',
                        fallback=cfg.get('auto_select_after_fetch', False)
                    )
                except Exception:
                    cfg['auto_select_after_fetch'] = cfg.get('auto_select_after_fetch', False)
                
                # 安全地读取整数
                try:
                    threads_str = cp['main'].get('threads', str(cfg.get('threads', DEFAULT_THREAD_COUNT)))
                    cfg['threads'] = int(threads_str)
                except Exception:
                    try:
                        cfg['threads'] = int(cfg.get('threads', DEFAULT_THREAD_COUNT))
                    except Exception:
                        cfg['threads'] = DEFAULT_THREAD_COUNT
            
            # 加载用户列表
            cfg['users'] = []
            if 'users' in cp:
                for key in cp['users']:
                    if key.startswith('user'):
                        try:
                            value = cp['users'][key]
                            parts = value.split(',', 1)
                            if len(parts) == 2:
                                cfg['users'].append({'username': parts[0].strip(), 'url': parts[1].strip()})
                        except Exception:
                            pass  # 跳过格式错误的用户
        except Exception:
            pass  # 静默失败

    # 确保关键默认值存在
    cfg.setdefault('path', '')
    cfg.setdefault('cookie', '')
    cfg.setdefault('use_mix_folder', True)
    cfg.setdefault('include_date_in_filename', True)
    cfg.setdefault('auto_select_after_fetch', False)
    cfg.setdefault('threads', DEFAULT_THREAD_COUNT)
    cfg.setdefault('users', [])
    
    return cfg


def save_config(cfg):
    """保存配置到INI文件"""
    try:
        cp = configparser.ConfigParser(interpolation=None)
        cp['main'] = {
            'path': cfg.get('path', ''),
            'use_mix_folder': str(bool(cfg.get('use_mix_folder', True))),
            'include_date_in_filename': str(bool(cfg.get('include_date_in_filename', True))),
            'auto_select_after_fetch': str(bool(cfg.get('auto_select_after_fetch', False))),
            'threads': str(int(cfg.get('threads', DEFAULT_THREAD_COUNT))),
            'cookie': cfg.get('cookie', ''),
        }
        
        # 保存用户列表
        if 'users' in cfg and cfg['users']:
            cp['users'] = {}
            for idx, user in enumerate(cfg['users'], start=1):
                cp['users'][f'user{idx}'] = f"{user.get('username', '')},{user.get('url', '')}"
        
        # 自定义写入，在每个配置项之间添加空行
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            for section in cp.sections():
                f.write(f'[{section}]\n')
                for key, value in cp.items(section):
                    f.write(f'{key} = {value}\n')
                    f.write('\n')  # 每个配置项后添加空行
    except Exception as e:
        print(f"[警告] 保存 {CONFIG_FILE} 失败: {e}")


# ========================================
# 抖音API - 用户信息和作品获取
# ========================================

def extract_sec_user_id_from_url(user_home_url):
    """从主页URL中提取sec_user_id"""
    patterns = [
        r'/user/([A-Za-z0-9_.\-]+)',      # 匹配 /user/SEC_ID
        r'sec_user_id=([A-Za-z0-9_.\-]+)', # 匹配 sec_user_id=SEC_ID
        r'(MS4wLjAB[A-Za-z0-9_-]+)'        # 匹配 MS4wLjAB... 格式的ID
    ]
    
    for pattern in patterns:
        m = re.search(pattern, user_home_url)
        if m:
            return m.group(1)
    
    return None


def resolve_short_url_and_extract(url, timeout=10):
    """解析短链接/分享链接并提取sec_user_id"""
    try:
        # allow_redirects=True 会自动处理 301/302 跳转
        r = requests.get(url, allow_redirects=True, timeout=timeout)
        final_url = r.url or url
    except Exception:
        final_url = url
    
    # 优先从跳转后的URL提取
    sec = extract_sec_user_id_from_url(final_url)
    if not sec and final_url != url:
        # 如果跳转后没取到，尝试从原始URL取（防止跳转到登录页等）
        sec = extract_sec_user_id_from_url(url)
    
    return sec


def get_user_profile_info(session, sec_user_id):
    """获取用户资料信息（昵称、作品数）"""
    api_url = (
        f"https://www.douyin.com/aweme/v1/web/user/profile/other/"
        f"?device_platform=webapp&aid=6383&channel=channel_pc_web"
        f"&sec_user_id={sec_user_id}&from_user_page=1"
    )
    
    try:
        # 使用传入的 session 发起请求
        r = session.get(api_url, timeout=REQUEST_TIMEOUT)
        data = r.json()
        
        if data.get('status_code') == 0 and 'user' in data:
            u = data['user']
            return {
                'nickname': u.get('nickname') or '',
                'aweme_count': u.get('aweme_count', None) # 作品数
            }
        
        return None
    except Exception:
        # 异常（如JSON解析失败、超时）
        print(TEXT_WARN_PROFILE_FAIL)
        return None

# =============================================
# 作品解析与任务构建区块 （解析 Aweme → 视频/图片任务）
# =============================================

def get_second_highest_bitrate_url(aweme):
    """
    从 aweme JSON 对象中获取第二高码率的视频链接
    """
    try:
        video_info = aweme.get('video', {}) or {}
        bit_rate_list = video_info.get('bit_rate', []) or []
        
        if len(bit_rate_list) >= 2:
            # 按码率排序
            sorted_rates = sorted(bit_rate_list, key=lambda x: x.get('bit_rate', 0), reverse=True)
            # 获取第二高码率
            second_best = sorted_rates[1]
            url_list = second_best.get('play_addr', {}).get('url_list', [])
            if url_list:
                return url_list[0]  # 返回第一个链接
    except Exception:
        pass
    return None

def extract_media_links_from_aweme(aweme):
    """
    从单个 aweme JSON 对象中提取媒体链接。
    
    互斥提取逻辑：
      - 如果 image 项包含 'video' 字段 -> 视为实况图（只提取实况图视频 .mp4）
      - 否则 -> 视为普通图片（提取最高分辨率 url_list[-1] .jpg/.png）
      
    返回： desc, videos[], images[], live_images[], date_str, mix_name
    """
    videos, images, live_images = [], [], []
    aweme_id = aweme.get('aweme_id') or ''
    desc = aweme.get('desc', '') or aweme_id or 'no_desc'
    
    # 提取合集名称
    mix_name = None
    mix_info = aweme.get('mix_info') or {}
    if isinstance(mix_info, dict):
        mix_name = mix_info.get('mix_name') or mix_info.get('mix_name_str') or None
    if not mix_name:
        mix_name = aweme.get('mix_name') or aweme.get('mix_name_str') or None

    # 转换时间戳 -> YYYY-MM-DD
    ts = aweme.get('create_time')
    date_str = ''
    if ts:
        try:
            date_str = datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
        except Exception:
            pass # 时间戳转换失败

    # 自动截断过长描述
    if len(desc) > MAX_DESC_LENGTH:
        desc = desc[:MAX_DESC_LENGTH] + "......"

    # 1. 提取普通视频作品 (aweme.video)
    video_info = aweme.get('video', {}) or {}
    bit_rate_list = video_info.get('bit_rate', []) or []
    if bit_rate_list:
        try:
            # 选择最高码率
            best = max(bit_rate_list, key=lambda x: x.get('bit_rate', 0))
            url_list = best.get('play_addr', {}).get('url_list', [])
            if url_list:
                videos.append(url_list[0]) # 通常第一个链接最稳定
        except Exception:
            pass # 码率列表格式异常

    # 2. 提取图集作品 (aweme.images)
    if 'images' in aweme and isinstance(aweme['images'], list):
        for img in aweme['images']:
            if not isinstance(img, dict):
                continue

            # 2a. 检查是否为实况图（包含 video 字段）
            vinfo = img.get('video')
            if vinfo and isinstance(vinfo, dict) and 'bit_rate' in vinfo:
                try:
                    rates = vinfo.get('bit_rate') or []
                    if rates:
                        best = max(rates, key=lambda x: x.get('bit_rate', 0))
                        vurl_list = best.get('play_addr', {}).get('url_list', [])
                        if vurl_list:
                            live_images.append(vurl_list[0])
                    # 互斥：提取了实况图，就跳过该项的普通图片提取
                    continue
                except Exception:
                    pass # 码率列表格式异常

            # 2b. 按普通图片处理
            url_list = img.get('url_list', [])
            if url_list and isinstance(url_list, list) and url_list:
                # 默认最后一个是最高分辨率
                images.append(url_list[-1])

    return desc, videos, images, live_images, date_str, mix_name


def parse_all_awemes_to_tasks(all_awemes):
    """将所有aweme解析为下载任务列表"""
    video_tasks, image_tasks = [], []
    album_count = 0
    image_count = 0
    live_count = 0

    for aweme in all_awemes:
        desc, videos, images, live_images, date_str, mix_name = extract_media_links_from_aweme(aweme)

        # 视频任务
        for vurl in videos:
            ext = get_extension_from_url(vurl, '.mp4')
            video_tasks.append({'url': vurl, 'desc': desc, 'ext': ext, 'date': date_str, 'mix_name': mix_name, 'aweme': aweme})

        # 如果这个 aweme 有普通图片或实况图，则视为一个图集作品
        if images or live_images:
            album_count += 1

        # 普通图片（按张）
        for idx, iurl in enumerate(images, start=1):
            ext = get_extension_from_url(iurl, '.jpg')
            image_tasks.append({'url': iurl, 'desc': f"{desc}_p{idx}", 'ext': ext, 'date': date_str, 'mix_name': mix_name})
        image_count += len(images)

        # 实况图（按张），也放到 image_tasks，但 ext 可能是 .mp4
        for idx, lvurl in enumerate(live_images, start=1):
            ext = get_extension_from_url(lvurl, '.mp4')
            image_tasks.append({'url': lvurl, 'desc': f"{desc}_live{idx}", 'ext': ext, 'date': date_str, 'mix_name': mix_name})
        live_count += len(live_images)

    return video_tasks, image_tasks, album_count, image_count, live_count


# ========================================
# 下载引擎 - 多线程下载执行
# ========================================

def download_single_file(task, base_folder, is_image=False, worker=None, session=None):
    """
    下载单个文件（视频或图片/实况图）。
    由 `Worker.download_tasks` 在线程池中调用。
    """
    url = task['url']
    desc = task['desc']
    ext = task['ext']
    mix_name = task.get('mix_name') or None # None 表示不使用合集
    
    include_date = task.get('include_date_in_filename', True) # 默认为True，匹配旧行为
    date_str = task.get('date', '')
    
    base_filename = desc
    if include_date and date_str:
        base_filename = f"{date_str}_{desc}"

    # 1. 确定目标文件夹
    # 规则：base_folder / [mix_name?] / [images?]
    folder = base_folder
    if mix_name:
        mix_clean = sanitize_filename(mix_name, max_length=100)
        folder = os.path.join(folder, mix_clean)
    
    if is_image:
        # 图片/实况图 统一放入 images 子目录
        folder = os.path.join(folder, 'images')
        
    # 创建目标目录
    safe_mkdir(folder)

    # 2. 生成唯一文件名 (使用新的 base_filename)
    path = generate_unique_filename(base_filename, ext, folder, url)
    
    # 3. 执行下载
    headers = {'User-Agent': USER_AGENT, 'Referer': 'https://www.douyin.com/'}
    
    s = session
    if not s:
        s = requests.Session()
        s.headers.update(headers)
    
    try:
        # 使用 session.get()
        with s.get(url, headers=headers, stream=True, timeout=30) as r:
            r.raise_for_status()
            with open(path, 'wb') as f:
                # 分块下载，每块后检查是否需要停止
                for chunk in r.iter_content(8192):
                    if chunk:  # 过滤掉keep-alive新块
                        f.write(chunk)
                        # 如果提供了worker对象，检查是否需要停止下载
                        if worker and worker.should_stop_download():
                            # 中断下载并删除部分下载的文件
                            raise Exception("下载被用户终止")
        # 成功，返回相对路径
        relative_name = os.path.relpath(path, base_folder)
        return relative_name
    except Exception as e:
        # 失败（超时、HTTP错误、磁盘写入错误等）
        # 尝试删除可能已创建的不完整文件
        try:
            if os.path.exists(path):
                os.remove(path)
        except:
            pass
        
        if worker and worker.should_stop_download():
            return None
        if "下载被用户终止" in str(e): # 确保万无一失
             return None
        return None

# ========================================
# GUI界面 - PyQt6图形界面
# ========================================

def run_gui():
    """启动PyQt6图形界面"""
    try:
        from PyQt6 import QtWidgets, QtCore, QtGui
        from PyQt6.QtCore import Qt
    except Exception as e:
        print(f"[错误] PyQt6 未安装或无法导入: {e}\n请安装 PyQt6 后重试（pip install PyQt6）。")
        return

    # 启动时加载配置
    cfg = load_config()
    
    class NoFocusRectStyle(QtWidgets.QProxyStyle):
        """自定义样式类，用于禁用列表/树的焦点虚线框"""
        def drawPrimitive(self, element, option, painter, widget=None):
            # 禁止绘制焦点矩形
            if element == QtWidgets.QStyle.PrimitiveElement.PE_FrameFocusRect:
                return
            super().drawPrimitive(element, option, painter, widget)

    class Worker(QtCore.QObject):
        """
        后台工作线程（用于 Fetch 和 Download）
        """
        # 信号定义
        log_signal = QtCore.pyqtSignal(str) # 日志
        progress_signal = QtCore.pyqtSignal(int, int) # 进度 (done, total)
        tasks_signal = QtCore.pyqtSignal(object, object, object, object) # 任务 (vtasks, itasks, nickname, aweme_list)
        fetch_finished = QtCore.pyqtSignal() # 获取完成
        download_finished = QtCore.pyqtSignal() # 下载完成
        export_finished_signal = QtCore.pyqtSignal(str)  # Excel导出完成信号
        export_error_signal = QtCore.pyqtSignal(str)     # Excel导出错误信号
        finished = QtCore.pyqtSignal() # 线程退出

        def __init__(self, parent=None):
            super().__init__(parent)
            # 状态标志
            self._pause_requested = False
            self._fetch_stop_requested = False
            self._download_stop_requested = False
            
            # 结果存储
            self._failed_tasks = []
            self._completed_tasks = []
            self._total_received = 0
            self.all_awemes = []  # 存储所有获取到的aweme数据
            self.session = requests.Session()
            self.session.headers.update({'User-Agent': USER_AGENT})
            self._log_buffer = []
            self._log_buffer_lock = threading.Lock()
            
        def should_stop_download(self):
            """检查是否应该停止下载"""
            return getattr(self, '_download_stop_requested', False)
        
        def is_download_stopped(self):
            """检查下载是否已被用户停止"""
            return getattr(self, '_download_stop_requested', False)

        def fetch_tasks(self, url, cookie):
            """
            获取用户作品列表（在单独线程中运行）。
            采用分页增量方式，每获取一页就通过 tasks_signal 发回 GUI。
            """
            try:
                headers = {'Cookie': cookie, 'Referer': url}
                self.session.headers.update(headers)
                
                self.log_signal.emit('[信息] 开始获取用户信息')
                
                # 1. 解析 sec_user_id
                sec = resolve_short_url_and_extract(url)
                if not sec:
                    self.log_signal.emit('[错误] 无法解析 sec_user_id')
                    self.finished.emit()
                    return

                profile = get_user_profile_info(self.session, sec)
                if not profile:
                    self.log_signal.emit('[错误] 获取用户信息失败，Cookie 可能无效')
                    self.finished.emit()
                    return
                nickname = profile.get('nickname', '') or ''
                self.log_signal.emit(f"[信息] 抖音用户: {nickname}")

                # 3. 分页获取作品
                page = 1
                max_cursor = 0
                self._total_received = 0 # 重置计数
                
                # 用于批量发送任务的缓冲区
                batch_vtasks = []
                batch_itasks = []
                BATCH_SIZE = 50  # 每50个任务发送一次
                
                # 初始化aweme_list以确保变量始终定义
                aweme_list = []
                
                while True:
                    # 检查是否请求停止
                    if getattr(self, '_fetch_stop_requested', False):
                        self.log_signal.emit('[信息] 获取已停止')
                        break
                    
                    # 构造请求URL
                    if page == 1:
                        # 第一页
                        req_url = (
                            "https://www.douyin.com/aweme/v1/web/aweme/post/"
                            f"?device_platform=webapp&aid=6383&channel=channel_pc_web"
                            f"&sec_user_id={sec}&max_cursor=0&count={PAGE_COUNT_PER_REQUEST}"
                            f"&locate_query=false&show_live_replay_strategy=1&need_time_list=1"
                            f"&publish_video_strategy_type=2&from_user_page=1&update_version_code=170400"
                        )
                    else:
                        # 后续页
                        req_url = (
                            f"https://www-hj.douyin.com/aweme/v1/web/aweme/post/"
                            f"?device_platform=webapp&aid=6383&channel=channel_pc_web"
                            f"&sec_user_id={sec}&max_cursor={max_cursor}"
                            f"&count={PAGE_COUNT_PER_REQUEST}&locate_query=false"
                            f"&show_live_replay_strategy=1&need_time_list=0"
                            f"&publish_video_strategy_type=2&from_user_page=1&update_version_code=170400"
                        )
                    
                    # 发起请求
                    try:
                        r = self.session.get(req_url, timeout=REQUEST_TIMEOUT)
                        r.raise_for_status()
                        data = r.json()
                    except Exception as e:
                        self.log_signal.emit(f"[警告] 第 {page} 页请求异常: {e}")
                        break # 请求失败则终止后续获取

                    # 解析本页返回的 aweme 列表
                    aweme_list = data.get('aweme_list', []) or []
                    if not aweme_list:
                        break # 没有更多作品
                    
                    # 解析为任务
                    vtasks, itasks, _, _, _ = parse_all_awemes_to_tasks(aweme_list)
                    
                    # 添加到批量缓冲区
                    batch_vtasks.extend(vtasks)
                    batch_itasks.extend(itasks)
                    
                    # 累积aweme数据
                    if not hasattr(self, 'all_awemes'): 
                        self.all_awemes = []
                    self.all_awemes.extend(aweme_list)
                    
                    # 如果缓冲区达到批量大小或者这是最后一页，则发送任务
                    if len(batch_vtasks) + len(batch_itasks) >= BATCH_SIZE:
                        # (批量) 将任务发回 GUI
                        try:
                            self.tasks_signal.emit(batch_vtasks, batch_itasks, nickname, aweme_list)
                            batch_vtasks = []
                            batch_itasks = []
                        except Exception as e:
                            self.log_signal.emit(f"[警告] tasks_signal.emit 失败: {e}")
                    
                    # 更新累计计数并输出日志
                    self._total_received += len(aweme_list)
                    self.log_signal.emit(TEXT_INFO_FETCH_PAGE.format(page=page, count=len(aweme_list), total=self._total_received))
                    
                    # 准备下一页
                    max_cursor = data.get('max_cursor', 0)
                    has_more = data.get('has_more', 0) == 1
                    page += 1
                    time.sleep(DELAY_BETWEEN_PAGES)
                    
                    if not has_more:
                        break # 明确告知没有更多了
                
                # 发送剩余的任务
                if batch_vtasks or batch_itasks:
                    try:
                        # 使用空列表作为aweme_list的默认值，避免变量未定义
                        current_aweme_list = aweme_list if 'aweme_list' in locals() else []
                        self.tasks_signal.emit(batch_vtasks, batch_itasks, nickname, current_aweme_list)
                    except Exception as e:
                        self.log_signal.emit(f"[警告] tasks_signal.emit 失败: {e}")
                
                self.log_signal.emit('[完成] 获取完成')
                
            except Exception as e:
                self.log_signal.emit(f"[错误] 获取异常: {e}")
            finally:
                try:
                    # 告知界面 fetch 流程已完成
                    self.fetch_finished.emit()
                except Exception:
                    pass
                # 告知 GUI 线程已结束
                self.finished.emit()

        def _download_with_retry(self, task, base_folder, is_image, max_retries, session):
            """
            带重试机制的下载函数
            """
            # 检查是否是视频任务且包含aweme数据
            is_video_task = not is_image and 'aweme' in task
            
            for attempt in range(max_retries + 1):
                if self.should_stop_download():
                    return "__STOPPED__" # 返回特殊停止标识
                
                try:
                    result = download_single_file(task, base_folder, is_image, self, session)
                    if result:
                        # 下载成功
                        if attempt > 0:
                            # 如果是重试成功的，记录一下
                            print(f"[重试成功] {task['desc']} (尝试 {attempt + 1} 次)")
                        return result
                    
                    if self.should_stop_download():
                         return "__STOPPED__"

                    if attempt < max_retries:
                        # 下载失败但还有重试机会
                        # 如果是视频任务且是第一次失败，尝试使用第二高码率链接
                        if is_video_task and attempt == 0:
                            second_rate_url = get_second_highest_bitrate_url(task['aweme'])
                            if second_rate_url:
                                # 更新任务URL为第二高码率链接
                                task['url'] = second_rate_url
                                self.log_signal.emit(f"[信息] {task['desc']} 尝试使用第二高码率链接")
                                continue  # 直接重试，不等待
                        
                        time.sleep(2 ** attempt)  # 指数退避
                        if self.should_stop_download():
                            return "__STOPPED__"
                except Exception as e:
                    if "下载被用户终止" in str(e):
                        return "__STOPPED__"
                    
                    if attempt < max_retries:
                        # 如果是视频任务且是第一次失败，尝试使用第二高码率链接
                        if is_video_task and attempt == 0:
                            second_rate_url = get_second_highest_bitrate_url(task['aweme'])
                            if second_rate_url:
                                # 更新任务URL为第二高码率链接
                                task['url'] = second_rate_url
                                self.log_signal.emit(f"[信息] {task['desc']} 尝试使用第二高码率链接")
                                continue  # 直接重试，不等待
                        
                        time.sleep(2 ** attempt)  # 指数退避
                        if self.should_stop_download():
                            return "__STOPPED__"
                    else:
                        # 最后一次尝试仍然失败
                        # 不再 raise，而是返回 None
                        return None 
            return None # 所有重试失败

        def download_tasks(self, vtasks, itasks, base_folder, threads):
            """
            执行下载任务（在单独线程中运行）。
            使用线程池并发下载。
            """
            try:
                with self._log_buffer_lock:
                    self._log_buffer.clear()

                # 1. 过滤已存在的文件（断点续传）
                self.log_signal.emit('[信息] 检查已存在文件...')
                all_tasks = []
                results_success_files = set()
                
                # 检查视频
                for t in vtasks:
                    if getattr(self, '_download_stop_requested', False):
                        self.log_signal.emit('[信息] 下载任务已被用户终止')
                        return
                    
                    include_date = t.get('include_date_in_filename', True)
                    date_str = t.get('date', '')
                    expected = build_expected_filename(t['desc'], t['ext'], False, t.get('mix_name'), date_str, include_date)
                    
                    fullpath = os.path.join(base_folder, expected)
                    if os.path.exists(fullpath):
                        self.log_signal.emit(f"[跳过] 已存在: {expected}")
                        results_success_files.add(expected)
                        rec = {'task': t, 'is_image': False, 'path': expected}
                        self._completed_tasks.append(rec)
                    else:
                        all_tasks.append((t, False)) # (task, is_image=False)
                
                # 检查图片/实况
                for t in itasks:
                    if getattr(self, '_download_stop_requested', False):
                        self.log_signal.emit('[信息] 下载任务已被用户终止')
                        return
                    
                    include_date = t.get('include_date_in_filename', True)
                    date_str = t.get('date', '')
                    expected = build_expected_filename(t['desc'], t['ext'], True, t.get('mix_name'), date_str, include_date)

                    fullpath = os.path.join(base_folder, expected)
                    if os.path.exists(fullpath):
                        self.log_signal.emit(f"[跳过] 已存在: {expected}")
                        results_success_files.add(expected)
                        rec = {'task': t, 'is_image': True, 'path': expected}
                        self._completed_tasks.append(rec)
                    else:
                        all_tasks.append((t, True)) # (task, is_image=True)

                total = len(all_tasks)
                done = 0
                
                if total == 0:
                    self.log_signal.emit('[信息] 没有需要下载的新文件。')
                    self.progress_signal.emit(1, 1) # 进度100%
                    self.download_finished.emit()
                    self.finished.emit()
                    return

                # 2. 执行下载
                from concurrent.futures import ThreadPoolExecutor, as_completed
                
                # 下载重试次数
                MAX_RETRIES = 3
                
                with ThreadPoolExecutor(max_workers=threads) as ex:
                    future_map = {}
                    submitted_futures = []
                    
                    # 提交任务
                    for t, is_img in all_tasks:
                        # 在提交任务前检查是否需要停止
                        if getattr(self, '_download_stop_requested', False):
                            self.log_signal.emit('[信息] 下载任务已被用户终止')
                            return
                        
                        # 协作式暂停：在提交新任务前等待
                        while getattr(self, '_pause_requested', False):
                            if getattr(self, '_download_stop_requested', False):
                                self.log_signal.emit('[信息] 下载任务已被用户终止')
                                return
                            time.sleep(0.1)
                        
                        # 在提交任务前再次检查是否需要停止
                        if getattr(self, '_download_stop_requested', False):
                            self.log_signal.emit('[信息] 下载任务已被用户终止')
                            return
                            
                        # 在提交任务前再次检查是否需要停止
                        if getattr(self, '_download_stop_requested', False):
                            self.log_signal.emit('[信息] 下载任务已被用户终止')
                            return
                        
                        # 在提交任务前再次检查是否需要停止
                        if getattr(self, '_download_stop_requested', False):
                            self.log_signal.emit('[信息] 下载任务已被用户终止')
                            return
                        
                        future = ex.submit(self._download_with_retry, t, base_folder, is_img, MAX_RETRIES, self.session)
                        future_map[future] = (t, is_img)
                        submitted_futures.append(future)
                    
                    # 收集结果
                    # 创建一个完成任务的计数器
                    completed_count = 0
                    for future in as_completed(submitted_futures):
                        completed_count += 1
                        # 每处理一定数量的任务后检查是否需要停止
                        if completed_count % 5 == 0:  # 每5个任务检查一次
                            if self.should_stop_download():
                                self.log_signal.emit('[信息] 下载任务已被用户终止')
                                return
                        
                        t, is_img = future_map[future]
                        try:
                            # 获取结果前再次检查是否需要停止
                            if self.should_stop_download():
                                self.log_signal.emit('[信息] 下载任务已被用户终止')
                                return
                            result = future.result()
                            # 在处理结果前再次检查是否需要停止
                            if self.should_stop_download():
                                self.log_signal.emit('[信息] 下载任务已被用户终止')
                                return
                            
                            if result == "__STOPPED__":
                                # 用户主动停止，不计为失败
                                pass
                            elif result:
                                # 记录成功下载的文件
                                results_success_files.add(result)
                                rec = {'task': t, 'is_image': is_img, 'path': result}
                                self._completed_tasks.append(rec)
                                done += 1

                                logs_to_send = None
                                with self._log_buffer_lock:
                                    self._log_buffer.append(f"[完成] {result}")
                                    if len(self._log_buffer) >= 20: # 批量阈值
                                        logs_to_send = "\n".join(self._log_buffer)
                                        self._log_buffer.clear()
                                if logs_to_send:
                                    self.log_signal.emit(logs_to_send)
                                
                                # 优化：减少进度条更新频率
                                if done % 5 == 1 or done == total:
                                    self.progress_signal.emit(done, total)
                                    # 检查是否需要停止
                                    if self.should_stop_download():
                                        self.log_signal.emit('[信息] 下载任务已被用户终止')
                                        return
                            else:
                                # 真正失败 (result is None)
                                self.log_signal.emit(f"[失败] {t['desc']} - URL: {t['url']}")
                                self._failed_tasks.append(t)
                            
                            # 每处理一个任务后都检查是否需要停止
                            if self.should_stop_download():
                                self.log_signal.emit('[信息] 下载任务已被用户终止')
                                return
                        except Exception as e:
                            # 在处理异常前检查是否需要停止
                            if self.should_stop_download():
                                self.log_signal.emit('[信息] 下载任务已被用户终止')
                                return
                            self.log_signal.emit(f"[失败] {t['desc']} - URL: {t['url']} ({e})")
                            self._failed_tasks.append(t)
                
                # 确保最后一次进度更新
                self.progress_signal.emit(total, total)

                logs_to_send = None
                with self._log_buffer_lock:
                    if self._log_buffer:
                        logs_to_send = "\n".join(self._log_buffer)
                        self._log_buffer.clear()
                if logs_to_send:
                    self.log_signal.emit(logs_to_send)
                
                # 4. 完成
                # 标准化路径分隔符，确保日志中显示的路径斜杠一致
                normalized_base_folder = base_folder.replace('\\', '/').replace('\\', '/')
                self.log_signal.emit(f"[日志] 本次成功下载文件 {len(results_success_files)} 个（目录: {normalized_base_folder}）")
                
                # 仅当 download_tasks 正常结束时发出下载完成信号
                self.download_finished.emit()
                
                # 最终检查是否需要停止
                if self.should_stop_download():
                    self.log_signal.emit('[信息] 下载任务已被用户终止')
                    return

            except Exception as e:
                # 在处理异常时检查是否需要停止
                if self.should_stop_download():
                    self.log_signal.emit('[信息] 下载任务已被用户终止')
                    return
                self.log_signal.emit(f"[错误] 下载异常: {e}")
            finally:
                # 如果是用户主动停止下载，发出下载完成信号以更新UI
                if self.should_stop_download():
                    try:
                        self.download_finished.emit()
                    except:
                        pass
                # 告知 GUI 线程已结束
                self.finished.emit()
        
        def export_excel(self, all_awemes, nickname, base_folder):
            """在后台线程中执行Excel导出"""
            try:
                if not OPENPYXL_AVAILABLE:
                    self.export_error_signal.emit('[错误] 未安装openpyxl库，请运行: pip install openpyxl')
                    return

                # 创建Excel文件夹
                excel_folder = os.path.join(base_folder, '作品数据Excel')
                if not safe_mkdir(excel_folder):
                    self.export_error_signal.emit('[错误] 无法创建Excel文件夹')
                    return
                
                # 生成文件路径
                filename = f"{sanitize_filename(nickname)}.xlsx"
                filepath = os.path.join(excel_folder, filename)

                wb = openpyxl.Workbook()
                ws = wb.active
                if ws is not None:
                    ws.title = "作品数据"
                    
                    # 写入表头
                    headers = ['类型', '发布时间', '文案', '合集', '点赞数', '评论数', '收藏数', '分享数', '推荐次数', '视频时长', '作品链接']
                    ws.append(headers)
                    
                    # 写入数据
                    for aweme in all_awemes:
                        # 获取统计数据
                        statistics = aweme.get('statistics', {})
                        
                        # 获取类型
                        aweme_type = '视频'
                        if aweme.get('images'):
                            aweme_type = '图集'
                        
                        # 获取发布时间
                        create_time = aweme.get('create_time', 0)
                        if create_time:
                            try:
                                publish_time = datetime.fromtimestamp(create_time).strftime('%Y-%m-%d %H:%M:%S')
                            except:
                                publish_time = ''
                        else:
                            publish_time = ''
                        
                        # 获取合集名称
                        mix_name = ''
                        mix_info = aweme.get('mix_info', {})
                        if isinstance(mix_info, dict):
                            mix_name = mix_info.get('mix_name', '') or mix_info.get('mix_name_str', '') or ''
                        if not mix_name:
                            mix_name = aweme.get('mix_name', '') or aweme.get('mix_name_str', '') or ''
                        
                        # 获取视频时长
                        duration_text = ''
                        video_info = aweme.get('video', {})
                        if isinstance(video_info, dict):
                            duration = video_info.get('duration', 0)
                            if duration > 0:
                                # 转换毫秒为时分秒格式
                                total_seconds = duration // 1000
                                hours = total_seconds // 3600
                                minutes = (total_seconds % 3600) // 60
                                seconds = total_seconds % 60
                                
                                if hours > 0:
                                    duration_text = f"{hours}小时{minutes}分钟{seconds}秒"
                                elif minutes > 0:
                                    duration_text = f"{minutes}分钟{seconds}秒"
                                else:
                                    duration_text = f"{seconds}秒"
                        
                        aweme_id = aweme.get('aweme_id', '')
                        # 根据作品类型生成链接
                        if aweme_id:
                            if aweme.get('images'):
                                link = f"https://www.douyin.com/note/{aweme_id}"
                            else:
                                link = f"https://www.douyin.com/video/{aweme_id}"
                        else:
                            link = ''
                        
                        row_data = [
                            aweme_type,  # 类型
                            publish_time,  # 发布时间
                            aweme.get('desc', ''),  # 文案
                            mix_name,  # 合集
                            statistics.get('digg_count', 0),  # 点赞数
                            statistics.get('comment_count', 0),  # 评论数
                            statistics.get('collect_count', 0),  # 收藏数
                            statistics.get('share_count', 0),  # 分享数
                            statistics.get('recommend_count', 0),  # 推荐次数
                            duration_text,  # 视频时长
                            link  # 作品链接
                        ]
                        ws.append(row_data)
                    
                    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 对所有单元格应用左对齐样式和自动换行
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = alignment
                    
                    # 自动调整列宽，并为特定列设置最小宽度
                    for idx, column in enumerate(ws.columns, start=1):
                        max_length = 0
                        column_letter = get_column_letter(idx)
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        # 为特定列设置固定宽度
                        if idx == 1:  # 类型列
                            adjusted_width = 5  # 两个中文字宽度
                        elif idx == 4:  # 合集列
                            adjusted_width = 12  # 六个中文字宽度
                        elif idx == 9:  # 推荐次数列
                            adjusted_width = 9  # 四个中文字宽度
                        elif idx == 10:  # 视频时长列
                            adjusted_width = 12  # 六个中文字宽度
                        elif idx == 3:  # 文案列
                            # 文案列设置固定宽度以防止溢出
                            adjusted_width = 30  # 固定文案列宽度
                        else:
                            # 其他列根据内容自动调整宽度，但限制最大宽度
                            adjusted_width = min(max_length + 3, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                
                # 保存文件
                wb.save(filepath)
                
                # 发送完成信号
                self.export_finished_signal.emit(filepath)
                return
            except Exception as e:
                error_msg = f'[错误] 导出Excel失败: {str(e)}'
                self.export_error_signal.emit(error_msg)
                return
            
            # 错误情况
            error_msg = '[错误] 无法创建工作表'
            self.export_error_signal.emit(error_msg)


    class AboutWindow(QtWidgets.QDialog):
        """关于窗口"""
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setWindowTitle('关于')
            self.setModal(False)
            self.resize(600, 300)
            
            layout = QtWidgets.QVBoxLayout(self)
            
            self.about_text = QtWidgets.QTextEdit()
            self.about_text.setReadOnly(True)
            about_content = """
抖音主页作品解析下载 V3.4

作者:颜如嘤-YanRuYing

更新地址:https://www.52pojie.cn/thread-2064455-1-1.html

开源地址:https://github.com/yanruying/douyin-downloader

本项程序仅用于学习与研究
禁止将本程序用于任何商业或违法用途
            """
            self.about_text.setPlainText(about_content)
            layout.addWidget(self.about_text)
            
            btn_layout = QtWidgets.QHBoxLayout()
            self.close_btn = QtWidgets.QPushButton('关闭')
            btn_layout.addStretch()
            btn_layout.addWidget(self.close_btn)
            layout.addLayout(btn_layout)
            
            self.close_btn.clicked.connect(self.close)
            
            self.setStyleSheet("""
                QDialog { background-color: #ffffff; }
                QTextEdit {
                    border: 1px solid #dcdfe6; background: #ffffff; color: #303133;
                    padding: 6px; border-radius: 0px; font-size: 15px;
                    font-family: Consolas, 'Courier New', monospace;
                }
                QPushButton {
                    background-color: #409EFF; border: 1px solid #409EFF; color: white;
                    padding: 6px 14px; border-radius: 0px; font-weight: 500; font-size: 13px;
                }
                QPushButton:hover { background-color: #66b1ff; border: 1px solid #66b1ff; }
                QPushButton:pressed { background-color: #3a8ee6; border: 1px solid #3a8ee6; }
            """)


    class TutorialWindow(QtWidgets.QDialog):
        """Cookie 教程窗口"""
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setWindowTitle('Cookie 获取教程')
            self.setModal(False)
            self.resize(800, 500)
            
            layout = QtWidgets.QVBoxLayout(self)
            
            self.tutorial_text = QtWidgets.QTextEdit()
            self.tutorial_text.setReadOnly(True)
            tutorial_content = """
【Cookie 获取教程】

1. 打开浏览器（推荐使用 Chrome 或 Edge）

2. 访问抖音网页版：https://www.douyin.com/?recommend=1

3. 登录你的抖音账号

4. 按 F12 打开开发者工具

5. 点击顶部的 "Network" 或 "网络" 标签

6. 刷新页面（F5）

7. 一般第一个就是 ?recommend=1 请求

8. 点击请求，查看请求标头（Request Headers）

9. 找到 "Cookie" 字段，复制全部

10. 将复制的 Cookie 粘贴到下面的 Cookie 输入框中

11. 点击保存按钮

【注意事项】
- Cookie 具有时效性，如果出现获取失败，请重新获取 Cookie
- 不要将 Cookie 分享给他人，避免账号被盗
            """
            self.tutorial_text.setPlainText(tutorial_content)
            layout.addWidget(self.tutorial_text)
            
            btn_layout = QtWidgets.QHBoxLayout()
            self.close_btn = QtWidgets.QPushButton('关闭')
            btn_layout.addStretch()
            btn_layout.addWidget(self.close_btn)
            layout.addLayout(btn_layout)
            
            self.close_btn.clicked.connect(self.close)
            
            self.setStyleSheet("""
                QDialog { background-color: #ffffff; }
                QTextEdit {
                    border: 1px solid #dcdfe6; background: #ffffff; color: #303133;
                    padding: 6px; border-radius: 0px; font-size: 15px;
                    font-family: Consolas, 'Courier New', monospace;
                }
                QPushButton {
                    background-color: #409EFF; border: 1px solid #409EFF; color: white;
                    padding: 6px 14px; border-radius: 0px; font-weight: 500; font-size: 13px;
                }
                QPushButton:hover { background-color: #66b1ff; border: 1px solid #66b1ff; }
                QPushButton:pressed { background-color: #3a8ee6; border: 1px solid #3a8ee6; }
            """)


    class LogWindow(QtWidgets.QDialog):
        """日志窗口"""
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setWindowTitle('运行日志')
            self.setModal(False)
            self.resize(800, 500)
            
            layout = QtWidgets.QVBoxLayout(self)
            
            self.log_text = QtWidgets.QTextEdit()
            self.log_text.setReadOnly(True)
            layout.addWidget(self.log_text)
            
            btn_layout = QtWidgets.QHBoxLayout()
            self.clear_log_btn = QtWidgets.QPushButton('清空日志')
            self.close_btn = QtWidgets.QPushButton('关闭')
            btn_layout.addStretch()
            btn_layout.addWidget(self.clear_log_btn)
            btn_layout.addWidget(self.close_btn)
            layout.addLayout(btn_layout)
            
            self.clear_log_btn.clicked.connect(self.clear_log)
            self.close_btn.clicked.connect(self.close)
            
            self.setStyleSheet("""
                QDialog { background-color: #ffffff; }
                QTextEdit {
                    border: 1px solid #dcdfe6; background: #ffffff; color: #303133;
                    padding: 6px; border-radius: 0px; font-size: 15px;
                    font-family: Consolas, 'Courier New', monospace;
                }
                QPushButton {
                    background-color: #409EFF; border: 1px solid #409EFF; color: white;
                    padding: 6px 14px; border-radius: 0px; font-weight: 500; font-size: 13px;
                }
                QPushButton:hover { background-color: #66b1ff; border: 1px solid #66b1ff; }
                QPushButton:pressed { background-color: #3a8ee6; border: 1px solid #3a8ee6; }
            """)

        def append_log(self, text):
            self.log_text.append(text)
            # 自动滚动到底部
            self.log_text.moveCursor(QtGui.QTextCursor.MoveOperation.End)
        
        def clear_log(self):
            self.log_text.clear()


    class UserListWindow(QtWidgets.QDialog):
        """用户列表窗口"""
        
        def __init__(self, parent=None, checkmark_svg_path=''):
            super().__init__(parent)
            self.checkmark_svg_path = checkmark_svg_path
            self.setWindowTitle('主页链接')
            self.setModal(False)
            self.resize(800, 500)
            
            layout = QtWidgets.QVBoxLayout(self)
            
            # 用户列表表格
            self.user_tree = QtWidgets.QTreeWidget()
            self.user_tree.setStyle(NoFocusRectStyle()) # 禁用焦点虚线框
            self.user_tree.setHeaderLabels(['选择', '序号', '用户名', '主页链接', '操作'])
            self.user_tree.setRootIsDecorated(False)
            self.user_tree.setUniformRowHeights(False)
            self.user_tree.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
            self.user_tree.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
            self.user_tree.setAttribute(QtCore.Qt.WidgetAttribute.WA_MacShowFocusRect, False)
            
            # 设置列宽
            fm = self.user_tree.fontMetrics()
            width0 = fm.horizontalAdvance('选择') + 16
            col4_w = fm.horizontalAdvance('汉' * 4) + 12
            self.user_tree.setColumnWidth(0, width0)
            self.user_tree.setColumnWidth(1, 60)
            self.user_tree.setColumnWidth(2, 100)
            self.user_tree.setColumnWidth(4, int(col4_w))
            
            # 设置表头
            header = self.user_tree.header()
            if header:
                header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Fixed)
                header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Fixed)
                header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch) # 主页列弹性
                header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeMode.Fixed)
                header.setSectionsMovable(False)
                header.setStretchLastSection(False)
            
            layout.addWidget(self.user_tree)
            
            # 按钮区
            btn_layout = QtWidgets.QHBoxLayout()
            self.select_all_btn = QtWidgets.QPushButton('全选')
            self.delete_btn = QtWidgets.QPushButton('删除')
            self.delete_btn.setStyleSheet('''
                QPushButton {
                    background: #d9534f; color: white; padding: 7px 14px;
                    border: none; font-weight: 500; font-size: 13px;                 
                }
                QPushButton:hover { background: #fa8480; }
            ''')
            self.close_btn = QtWidgets.QPushButton('关闭')
            
            btn_layout.addWidget(self.select_all_btn)
            btn_layout.addWidget(self.delete_btn)
            btn_layout.addStretch()
            btn_layout.addWidget(self.close_btn)
            layout.addLayout(btn_layout)
            
            # 绑定事件
            self.select_all_btn.clicked.connect(self.on_select_all)
            self.delete_btn.clicked.connect(self.on_delete)
            self.close_btn.clicked.connect(self.close)
            self.user_tree.itemSelectionChanged.connect(self.on_selection_changed)
            
            # 设置窗口样式
            self.setStyleSheet("""
                QDialog { background-color: #ffffff; }
                QTreeWidget {
                    background: #ffffff; border: 1px solid #e4e7ed;
                    alternate-background-color: #fafbfc; gridline-color: #f2f6fc;
                    selection-background-color: #d9eaff; font-size: 13px;
                    show-decoration-selected: 0;
                }
                QTreeWidget::item { padding: 0px 4px; color: #222222; outline: 0; }
                QTreeWidget::item:focus { outline: 0; border: 0; }
                QTreeWidget::item:hover { background: #f3f8fe; }
                QTreeWidget::item:selected { background: #cfe4ff; color: #000000; }
                QTreeWidget::item:selected:active { background: #cfe4ff; outline: 0; }
                QTreeWidget::item:selected:!active { background: #cfe4ff; outline: 0; }
                QTreeWidget::indicator {
                    width: 16px; height: 16px; border: 1px solid #c0c4cc;
                    border-radius: 2px; background: #ffffff;
                }
                QTreeWidget::indicator:hover { border: 1px solid #409EFF; }
                QTreeWidget::indicator:checked {
                    background-color: #409EFF; border: 1px solid #409EFF;
                    image: url(""" + self.checkmark_svg_path + r""");
                }
                QPushButton {
                    background-color: #409EFF; border: 1px solid #409EFF; color: white;
                    padding: 6px 14px; border-radius: 0px; font-weight: 500; font-size: 13px;
                }
                QPushButton:hover { background-color: #66b1ff; border: 1px solid #66b1ff; }
                QPushButton:pressed { background-color: #3a8ee6; border: 1px solid #3a8ee6; }
            """)
            
            # 加载用户列表
            self.load_users()
        
        def load_users(self):
            """加载用户列表"""
            self.user_tree.clear()
            users = cfg.get('users', [])
            for idx, user in enumerate(users, start=1):
                item = QtWidgets.QTreeWidgetItem(self.user_tree, [' ', str(idx), user.get('username', ''), user.get('url', ''), ''])
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                item.setCheckState(0, Qt.CheckState.Unchecked)
                item.setTextAlignment(0, int(Qt.AlignmentFlag.AlignLeft))
                item.setTextAlignment(4, int(Qt.AlignmentFlag.AlignCenter))
                item.setData(0, Qt.ItemDataRole.UserRole, user)
                
                # 在操作列添加 "获取" 按钮
                fetch_btn = QtWidgets.QPushButton('获取')
                fm = self.user_tree.fontMetrics()
                btn_width = fm.horizontalAdvance('关闭') + 28
                fetch_btn.setFixedWidth(btn_width)
                fetch_btn.setFixedHeight(28)
                fetch_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #409EFF; border: 1px solid #409EFF; color: white;
                        padding: 0px; border-radius: 0px; font-weight: 500; font-size: 13px;
                    }
                    QPushButton:hover { background-color: #66b1ff; border: 1px solid #66b1ff; }
                    QPushButton:pressed { background-color: #3a8ee6; border: 1px solid #3a8ee6; }
                """)
                fetch_btn.clicked.connect(lambda checked, u=user: self.on_fetch_user(u))
                
                # 使用布局在单元格内居中按钮
                btn_container = QtWidgets.QWidget()
                btn_container.setMinimumHeight(50)
                v_layout = QtWidgets.QVBoxLayout(btn_container)
                v_layout.setContentsMargins(0, 0, 0, 0)
                v_layout.setSpacing(0)
                v_layout.addStretch()
                h_widget = QtWidgets.QWidget()
                h_layout = QtWidgets.QHBoxLayout(h_widget)
                h_layout.setContentsMargins(0, 0, 0, 0)
                h_layout.setSpacing(0)
                h_layout.addStretch()
                h_layout.addWidget(fetch_btn)
                h_layout.addStretch()
                v_layout.addWidget(h_widget)
                v_layout.addStretch()
                self.user_tree.setItemWidget(item, 4, btn_container)
                
                # 设置行高
                item.setSizeHint(0, QtCore.QSize(-1, 50))
        
        def on_fetch_user(self, user):
            """点击 "获取" 按钮"""
            try:
                # 获取主窗口
                main_window = self.parent()
                if main_window:
                    # 1. 填充主页链接
                    if hasattr(main_window, 'url_edit'):
                        url_edit = getattr(main_window, 'url_edit', None)
                        if url_edit:
                            url_edit.setText(user.get('url', ''))
                    # 2. 触发主窗口的 "获取"
                    if hasattr(main_window, 'on_fetch'):
                        on_fetch = getattr(main_window, 'on_fetch', None)
                        if on_fetch:
                            on_fetch()
                    # 3. 关闭弹窗
                    self.close()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, '错误', f'获取失败: {e}')
        
        def on_delete(self):
            """删除选中的用户"""
            selected_items = []
            for i in range(self.user_tree.topLevelItemCount()):
                item = self.user_tree.topLevelItem(i)
                if item and item.checkState(0) == Qt.CheckState.Checked:
                    selected_items.append(item)
            
            if not selected_items:
                QtWidgets.QMessageBox.warning(self, '提示', '请先选择要删除的用户')
                return
            
            msg_box = QtWidgets.QMessageBox(self)
            msg_box.setWindowTitle('确认')
            msg_box.setText(f'确定要删除选中的 {len(selected_items)} 个用户吗？')
            msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Cancel)
            msg_box.setDefaultButton(QtWidgets.QMessageBox.StandardButton.Cancel)
            button_ok = msg_box.button(QtWidgets.QMessageBox.StandardButton.Ok)
            button_cancel = msg_box.button(QtWidgets.QMessageBox.StandardButton.Cancel)
            if button_ok: button_ok.setText('确认')
            if button_cancel: button_cancel.setText('取消')
            
            if msg_box.exec() != QtWidgets.QMessageBox.StandardButton.Ok:
                return
            
            # 从配置中删除
            users = cfg.get('users', [])
            users_to_remove = [item.data(0, Qt.ItemDataRole.UserRole) for item in selected_items]
            
            new_users = [u for u in users if u not in users_to_remove]
        
            cfg['users'] = new_users
            save_config(cfg)
            
            # 重新加载列表
            self.load_users()
        
        def on_select_all(self):
            """全选/反选所有用户"""
            # 检查是否有未选中的项
            has_unchecked = False
            for i in range(self.user_tree.topLevelItemCount()):
                item = self.user_tree.topLevelItem(i)
                if item and item.checkState(0) == Qt.CheckState.Unchecked:
                    has_unchecked = True
                    break
            
            # 如果有未选中的，则全选；否则全不选
            new_state = Qt.CheckState.Checked if has_unchecked else Qt.CheckState.Unchecked
            for i in range(self.user_tree.topLevelItemCount()):
                item = self.user_tree.topLevelItem(i)
                if item:
                    item.setCheckState(0, new_state)
        
        def on_selection_changed(self):
            """当选择改变时，同步复选框状态（行选 -> 勾选）"""
            selected_items = self.user_tree.selectedItems()
            # 遍历所有项，根据是否被选中来设置复选框
            for i in range(self.user_tree.topLevelItemCount()):
                item = self.user_tree.topLevelItem(i)
                if item:
                    if item in selected_items:
                        item.setCheckState(0, Qt.CheckState.Checked)
                    else:
                        item.setCheckState(0, Qt.CheckState.Unchecked)


    class SettingsWindow(QtWidgets.QDialog):
        """设置窗口"""
        def __init__(self, parent=None, checkmark_svg_path=''):
            super().__init__(parent)
            self.checkmark_svg_path = checkmark_svg_path
            self.setWindowTitle('设置')
            self.setModal(False)
            self.resize(500, 550)
            
            layout = QtWidgets.QVBoxLayout(self)
            layout.setSpacing(12)
            
            # 顶部按钮组
            top_btns_layout = QtWidgets.QHBoxLayout()
            self.view_log_btn = QtWidgets.QPushButton('查看日志')
            self.tutorial_btn = QtWidgets.QPushButton('查看教程')
            self.about_btn = QtWidgets.QPushButton('关于')
            top_btns_layout.addWidget(self.view_log_btn)
            top_btns_layout.addStretch()
            top_btns_layout.addWidget(self.tutorial_btn)
            top_btns_layout.addWidget(self.about_btn)
            layout.addLayout(top_btns_layout)
            layout.addSpacing(6)
            
            # Cookie
            layout.addWidget(QtWidgets.QLabel('Cookie:'))
            self.settings_cookie = QtWidgets.QTextEdit()
            self.settings_cookie.setPlainText(cfg.get('cookie', ''))
            self.settings_cookie.setFixedHeight(80)
            layout.addWidget(self.settings_cookie)
            layout.addSpacing(6)
            
            # 保存路径
            path_layout = QtWidgets.QHBoxLayout()
            path_layout.addWidget(QtWidgets.QLabel('保存路径:'))
            self.settings_path = QtWidgets.QLineEdit()
            self.settings_path.setText(cfg.get('path', ''))
            path_layout.addWidget(self.settings_path)
            self.settings_browse_btn = QtWidgets.QPushButton('浏览')
            path_layout.addWidget(self.settings_browse_btn)
            layout.addLayout(path_layout)
            layout.addSpacing(6)
            
            # 线程数
            threads_layout = QtWidgets.QHBoxLayout()
            threads_layout.addWidget(QtWidgets.QLabel('下载线程:'))
            self.threads_spin = QtWidgets.QSpinBox()
            self.threads_spin.setMinimum(1)
            self.threads_spin.setMaximum(64)
            try:
                self.threads_spin.setValue(int(cfg.get('threads', DEFAULT_THREAD_COUNT)))
            except Exception:
                self.threads_spin.setValue(DEFAULT_THREAD_COUNT)
            # 去掉上下箭头
            try:
                self.threads_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
            except Exception:
                pass
            threads_layout.addWidget(self.threads_spin)
            threads_layout.addStretch()
            layout.addLayout(threads_layout)
            layout.addSpacing(10)
            
            # 复选项
            self.chk_mix_setting = QtWidgets.QCheckBox('合集统一下载到【合集名称】文件夹')
            self.chk_mix_setting.setChecked(bool(cfg.get('use_mix_folder', True)))
            layout.addWidget(self.chk_mix_setting)
            layout.addSpacing(4)
            
            self.chk_date_setting = QtWidgets.QCheckBox('文件名前缀增加作品发布时间')
            self.chk_date_setting.setChecked(bool(cfg.get('include_date_in_filename', True)))
            layout.addWidget(self.chk_date_setting)
            layout.addSpacing(4)
            
            self.chk_auto_select = QtWidgets.QCheckBox('作品获取完成后自动全选')
            self.chk_auto_select.setChecked(bool(cfg.get('auto_select_after_fetch', True)))
            layout.addWidget(self.chk_auto_select)
            
            layout.addStretch()
            
            # 底部按钮
            button_layout = QtWidgets.QHBoxLayout()
            self.save_settings_btn = QtWidgets.QPushButton('保存')
            self.cancel_btn = QtWidgets.QPushButton('取消')
            button_layout.addWidget(self.save_settings_btn)
            button_layout.addWidget(self.cancel_btn)
            layout.addLayout(button_layout)
            
            # 绑定事件
            self.save_settings_btn.clicked.connect(self.save_settings)
            self.cancel_btn.clicked.connect(self.close)
            self.settings_browse_btn.clicked.connect(self.on_browse_path)
            self.about_btn.clicked.connect(self.on_view_about)
            self.view_log_btn.clicked.connect(self.on_view_log)
            self.tutorial_btn.clicked.connect(self.on_view_tutorial)
            
            # 初始化子窗口（但不显示）
            self.about_window = AboutWindow(self)
            self.about_window.hide()
            self.tutorial_window = TutorialWindow(self)
            self.tutorial_window.hide()
            
            # 设置样式
            self.setStyleSheet("""QDialog {
    background-color: #ffffff;
}
QLabel { color: #303133; font-size: 13px; }
QLineEdit {
    border: 1px solid #dcdfe6; background: #ffffff; color: #303133;
    padding: 6px; border-radius: 0px; font-size: 13px;
}
QLineEdit:focus { border: 1px solid #409EFF; background: #f9fcff; }
QTextEdit {
    border: 1px solid #dcdfe6; background: #ffffff; color: #303133;
    padding: 6px; border-radius: 0px; font-size: 13px;
}
QTextEdit:focus { border: 1px solid #409EFF; background: #f9fcff; }
QPushButton {
    border: 1px solid #dcdfe6; background: #409EFF; color: #ffffff;
    padding: 6px 12px; border-radius: 0px; font-size: 13px;
}
QPushButton:hover { background: #66b1ff; }
QPushButton:pressed {
    border: 1px solid #409EFF; background: #409EFF; color: #ffffff;
}
QCheckBox { color: #303133; font-size: 13px; spacing: 8px; }
QCheckBox::indicator {
    width: 16px; height: 16px; border: 1px solid #c0c4cc;
    border-radius: 2px; background: #ffffff;
}
QCheckBox::indicator:hover { border: 1px solid #409EFF; }
QCheckBox::indicator:checked {
    background-color: #409EFF; border: 1px solid #409EFF;
    image: url(""" + self.checkmark_svg_path + r""");
}
QSpinBox {
    border: 1px solid #dcdfe6; padding: 6px; background: #ffffff;
    color: #303133; border-radius: 0px; font-size: 13px;
}
QSpinBox:focus { border: 1px solid #409EFF; background: #f9fcff; }
/* 重写底部 Save/Cancel 按钮样式 */
QPushButton#save_settings_btn, QPushButton#cancel_btn {
    background-color: #409EFF; border: 1px solid #409EFF; color: white;
    padding: 6px 14px; border-radius: 0px; font-weight: 500; font-size: 13px;
}
QPushButton#save_settings_btn:hover, QPushButton#cancel_btn:hover {
    background-color: #66b1ff; border: 1px solid #66b1ff;
}
QPushButton#save_settings_btn:pressed, QPushButton#cancel_btn:pressed {
    background-color: #3a8ee6; border: 1px solid #3a8ee6;
}
""")
            # 为按钮设置对象名称以便 CSS 选择器生效
            self.save_settings_btn.setObjectName("save_settings_btn")
            self.cancel_btn.setObjectName("cancel_btn")
            
        
        def on_view_about(self):
            """显示关于窗口"""
            try:
                if self.about_window:
                    self.about_window.show()
                    self.about_window.raise_()
                    self.about_window.activateWindow()
            except Exception:
                pass
        
        def on_view_log(self):
            """显示日志窗口（从主窗口获取）"""
            try:
                parent_window = self.parent()
                if parent_window and hasattr(parent_window, 'log_window'):
                    log_window = getattr(parent_window, 'log_window', None)
                    if log_window:
                        log_window.show()
                        log_window.raise_()
                        log_window.activateWindow()
            except Exception:
                pass
        
        def on_view_tutorial(self):
            """显示教程窗口"""
            try:
                if self.tutorial_window:
                    self.tutorial_window.show()
                    self.tutorial_window.raise_()
                    self.tutorial_window.activateWindow()
            except Exception:
                pass
        
        def on_browse_path(self):
            """浏览选择保存路径"""
            dlg = QtWidgets.QFileDialog(self)
            p = dlg.getExistingDirectory(self, '选择目录', self.settings_path.text() or os.getcwd())
            if p:
                self.settings_path.setText(p)
        
        def save_settings(self):
            """保存设置"""
            # 1. 将设置写入全局 cfg 变量
            cfg['cookie'] = self.settings_cookie.toPlainText().strip()
            cfg['path'] = self.settings_path.text().strip()
            cfg['use_mix_folder'] = bool(self.chk_mix_setting.isChecked())
            cfg['include_date_in_filename'] = bool(self.chk_date_setting.isChecked())
            cfg['auto_select_after_fetch'] = bool(self.chk_auto_select.isChecked())
            cfg['threads'] = int(self.threads_spin.value())
            
            try:
                # 2. 持久化到 config.ini
                save_config(cfg)
                
                # 3. 通知主窗口
                parent_window = self.parent()
                if parent_window and hasattr(parent_window, 'append_log'):
                    append_log_func = getattr(parent_window, 'append_log', None)
                    if append_log_func:
                        append_log_func('[信息] 设置已保存')
                self.close()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, '警告', f'保存设置失败: {e}')


    class MainWindow(QtWidgets.QMainWindow):
        """主窗口"""
        def __init__(self, checkmark_svg_path=''):
            super().__init__()
            self.checkmark_svg_path = checkmark_svg_path
            self.setWindowTitle(TEXT_APP_NAME)
            self.resize(1200, 700)
            central = QtWidgets.QWidget()
            self.setCentralWidget(central)
            lay = QtWidgets.QVBoxLayout(central)

            # 1. 表单区（URL输入）
            form = QtWidgets.QGridLayout()
            lay.addLayout(form)
            # "主页链接" 标签做成按钮，点击可打开用户列表
            self.url_label_btn = QtWidgets.QPushButton('主页链接:')
            self.url_label_btn.setFlat(True)
            self.url_label_btn.setCursor(QtGui.QCursor(Qt.CursorShape.PointingHandCursor))
            form.addWidget(self.url_label_btn, 0, 0)
            self.url_edit = QtWidgets.QLineEdit()
            form.addWidget(self.url_edit, 0, 1, 1, 1)
            self.fetch_btn = QtWidgets.QPushButton('获取作品')
            form.addWidget(self.fetch_btn, 0, 2)


            # 2. 按钮区
            btns = QtWidgets.QHBoxLayout()
            lay.addLayout(btns)
            self.settings_btn = QtWidgets.QPushButton('设置')
            self.clear_btn = QtWidgets.QPushButton('清空列表')
            self.select_all_btn = QtWidgets.QPushButton('全选')
            self.export_excel_btn = QtWidgets.QPushButton('导出Excel')  # 新增导出Excel按钮
            self.invert_btn = QtWidgets.QPushButton('反选')
            self.download_btn = QtWidgets.QPushButton('开始下载')
            
            btns.addWidget(self.settings_btn)
            btns.addWidget(self.clear_btn)
            btns.addStretch()  # 弹性空间
            btns.addWidget(self.export_excel_btn)  # 添加导出按钮
            
            if not OPENPYXL_AVAILABLE:
                self.export_excel_btn.setEnabled(False)
                self.export_excel_btn.setToolTip("请先安装 'openpyxl' (pip install openpyxl) 以启用此功能")
            
            btns.addWidget(self.select_all_btn)
            btns.addWidget(self.invert_btn)
            btns.addWidget(self.download_btn)

            # 3. 列表区
            self.tree = QtWidgets.QTreeWidget()
            self.tree.setStyle(NoFocusRectStyle()) # 禁用焦点虚线框
            self.tree.setHeaderLabels(['选择', '序号', '发布日期', '描述', '合集', '类型'])
            
            # 设置列宽
            fm = self.tree.fontMetrics()
            width0 = fm.horizontalAdvance('选择') + 16
            col4_w = fm.horizontalAdvance('汉' * 4) + 12
            col5_w = fm.horizontalAdvance('汉' * 4) + 12
            self.tree.setColumnWidth(0, width0)     # 选择
            self.tree.setColumnWidth(1, 60)         # 序号
            self.tree.setColumnWidth(2, 100)        # 发布日期
            self.tree.setColumnWidth(3, 360)        # 描述 (动态)
            self.tree.setColumnWidth(4, int(col4_w))# 合集
            self.tree.setColumnWidth(5, int(col5_w))# 类型
            
            # 设置表头
            header = self.tree.header()
            hdr_h = fm.height() + 10
            if header:
                header.setFixedHeight(int(hdr_h))
                header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Fixed)
                header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
                header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch) # 描述列弹性
                header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeMode.Fixed)
                header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeMode.Fixed)
                header.setSectionsMovable(False)
                header.setStretchLastSection(False)
                
            self.tree.setRootIsDecorated(False)
            self.tree.setUniformRowHeights(True)
            self.tree.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
            self.tree.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
            self.tree.setAttribute(QtCore.Qt.WidgetAttribute.WA_MacShowFocusRect, False)
            self.tree.setFrameShape(QtWidgets.QFrame.Shape.Box)
            self.tree.setAlternatingRowColors(True) # 交替行颜色
            self.tree.setStyleSheet(
                "QTreeWidget { background: #ffffff; border: 1px solid #e6eef8; show-decoration-selected: 0; }"
                "QTreeWidget::item { padding:6px 4px; color: #222222; outline: 0; }"
                "QTreeWidget::item:focus { outline: 0; border: 0; }"
                "QTreeWidget::item:selected { background: #e6f2ff; color: #000000; outline: 0; }"
                "QTreeWidget::item:selected:active { background: #e6f2ff; outline: 0; }"
                "QTreeWidget::item:selected:!active { background: #e6f2ff; outline: 0; }"
            )
            lay.addWidget(self.tree)

            # 4. 进度与状态
            bottom = QtWidgets.QHBoxLayout()
            lay.addLayout(bottom)
            self.progress = QtWidgets.QProgressBar()
            self.progress.setFixedHeight(26)
            self.progress.setTextVisible(True)
            self.progress.setStyleSheet(
                "QProgressBar { border: none; border-radius: 0px; background: #f0f0f0; text-align: center; }"
                "QProgressBar::chunk { background-color: #5aa6ff; border-radius: 0px; }")
            bottom.addWidget(self.progress)
            self.progress.hide() # 默认隐藏
            
            # 状态栏
            status_layout = QtWidgets.QHBoxLayout()
            lay.addLayout(status_layout)
            self.status = QtWidgets.QLabel('')
            self.status.setCursor(QtGui.QCursor(Qt.CursorShape.PointingHandCursor)) # 手型
            self.status.setMouseTracking(True)
            status_layout.addWidget(self.status)
            status_layout.addStretch()
            status_layout.addWidget(QtWidgets.QLabel('当前用户:'))
            self.nickname_label = QtWidgets.QLabel('')
            font = self.nickname_label.font()
            font.setBold(True)
            self.nickname_label.setFont(font)
            status_layout.addWidget(self.nickname_label)

            # 5. 数据与状态
            self.vtasks_all = [] # 存储所有获取到的视频任务
            self.itasks_all = [] # 存储所有获取到的图片任务
            self.vtasks = []
            self.itasks = []
            self.all_awemes = []  # 存储所有获取到的aweme原始数据，用于导出Excel
            self.current_nickname = '' 

            # 6. 初始化子窗口（隐藏）
            self.log_window = LogWindow(self)
            self.log_window.hide()
            self.user_list_window = UserListWindow(self, self.checkmark_svg_path)
            self.user_list_window.hide()
            self.settings_window = SettingsWindow(self, self.checkmark_svg_path)
            self.settings_window.hide()

            # 7. 初始化 Worker 线程
            self.worker = Worker()
            self._thread = None

            # 8. 样式调整
            btn_font = QtGui.QFont()
            btn_font.setPointSize(11)
            for b in (self.fetch_btn, self.download_btn, self.settings_btn, self.clear_btn, self.select_all_btn, self.invert_btn):
                b.setFont(btn_font)
            button_width = 100
            self.fetch_btn.setFixedWidth(button_width)
            self.download_btn.setFixedWidth(button_width)
            
            # 保持 clear_btn 的红色样式不变
            self.clear_btn.setStyleSheet('''
                QPushButton {
                    background: #d9534f; color: white; padding: 7px 14px;
                    border: none; font-weight: 500; font-size: 13px;                 
                }
                QPushButton:hover { background: #fa8480; }
                QPushButton:disabled { background: #f0b3b3; color: #f8e6e6; }
            ''')

            # 9. 事件绑定
            self.url_label_btn.clicked.connect(self.on_show_user_list)
            self.fetch_btn.clicked.connect(self.on_fetch)
            self.download_btn.clicked.connect(self.on_download)
            self.settings_btn.clicked.connect(self.on_settings)
            self.select_all_btn.clicked.connect(self.on_select_all)
            self.export_excel_btn.clicked.connect(self.on_export_excel)  # 导出Excel按钮事件
            self.invert_btn.clicked.connect(self.on_invert)
            self.clear_btn.clicked.connect(self.on_clear_list)
            self.status.mousePressEvent = lambda ev: self.on_status_click(ev) # 状态栏点击

            # Worker 信号
            self.worker.log_signal.connect(self.append_log)
            self.worker.tasks_signal.connect(lambda vtasks, itasks, nickname, aweme_list: self.on_tasks_received(vtasks, itasks, nickname, aweme_list))
            self.worker.progress_signal.connect(self.on_progress)
            self.worker.finished.connect(self.on_worker_finished)
            self.worker.fetch_finished.connect(self.on_fetch_finished)
            self.worker.download_finished.connect(self.on_download_finished)
            self.worker.export_finished_signal.connect(self._on_export_finished)
            self.worker.export_error_signal.connect(self._on_export_error)

            # 列表选择/复选联动
            self.tree.itemSelectionChanged.connect(self.on_tree_selection_changed)
            self.tree.itemChanged.connect(self.on_tree_item_changed)

            # 状态标志
            self._programmatic_change = False # 防止联动循环
            self._last_status_text = '' # 状态栏基础文本
            
            # 首次运行检查
            if not os.path.exists(CONFIG_FILE):
                QtCore.QTimer.singleShot(500, self.show_first_time_settings)

        def append_log(self, text):
            """向日志窗口和状态栏输出日志"""
            ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            if '\n' in text:
                lines = text.split('\n')
                formatted_lines = [f"[{ts}] {line}" for line in lines]
                full_log_text = "\n".join(formatted_lines)
                # 状态栏只显示最后一条
                self._last_status_text = lines[-1]
            else:
                full_log_text = f"[{ts}] {text}"
                self._last_status_text = text

            # 1. 添加到日志窗口（即使隐藏）
            if self.log_window:
                self.log_window.append_log(full_log_text)
            
            # 2. 更新状态栏
            self.update_status_label()

        def update_status_label(self):
            """更新状态栏（基础文本 + 选择计数）"""
            try:
                count = 0
                for i in range(self.tree.topLevelItemCount()):
                    it = self.tree.topLevelItem(i)
                    if it and it.checkState(0) == Qt.CheckState.Checked:
                        count += 1
                base = getattr(self, '_last_status_text', '') or ''
                if count > 0:
                    self.status.setText(f"{base} （已选择 {count} 个）")
                else:
                    self.status.setText(base)
            except Exception:
                pass # 忽略更新失败

        def on_tree_selection_changed(self):
            """处理列表选择变化 (行选 -> 勾选)"""
            if getattr(self, '_programmatic_change', False):
                return
            self._programmatic_change = True
            try:
                for i in range(self.tree.topLevelItemCount()):
                    it = self.tree.topLevelItem(i)
                    if it:
                        if it.isSelected():
                            it.setCheckState(0, Qt.CheckState.Checked)
                        else:
                            it.setCheckState(0, Qt.CheckState.Unchecked)
            finally:
                self._programmatic_change = False
            self.update_status_label() # 更新选择计数

        def on_tree_item_changed(self, item, column):
            """处理复选框变化 (勾选 -> 行选)"""
            if getattr(self, '_programmatic_change', False):
                return
            self._programmatic_change = True
            try:
                if column == 0:
                    state = item.checkState(0)
                    if state == Qt.CheckState.Checked:
                        item.setSelected(True)
                    else:
                        item.setSelected(False)
            finally:
                self._programmatic_change = False
            self.update_status_label() # 更新选择计数

        def on_progress(self, done, total):
            """更新进度条"""
            if not self.progress.isVisible():
                self.progress.show()
                
            self.progress.setMaximum(total)
            self.progress.setValue(done)
            pct = int((done / max(1, total)) * 100)
            self.progress.setFormat(f"%v / %m ({pct}%)")
            
            # 完成时变绿
            try:
                if total > 0 and done >= total:
                    self.progress.setStyleSheet(
                        "QProgressBar { border: none; border-radius: 0px; background: #f0f0f0; text-align: center; }"
                        "QProgressBar::chunk { background-color: #4CC14C; border-radius: 0px; }"
                    )
                else:
                    # 恢复进行中颜色（蓝色）
                    self.progress.setStyleSheet(
                        "QProgressBar { border: none; border-radius: 0px; background: #f0f0f0; text-align: center; }"
                        "QProgressBar::chunk { background-color: #5aa6ff; border-radius: 0px; }"
                    )
            except Exception:
                pass
        
        def on_download_finished(self):
            """下载完成处理（确保进度条是绿色）"""
            try:
                maxv = self.progress.maximum() or self.progress.value() or 1
                self.progress.setValue(maxv)
                # 检查是否是用户主动停止下载
                if hasattr(self.worker, '_download_stop_requested') and self.worker._download_stop_requested:
                    self.progress.setFormat(f"%v / %m (已停止)")
                    # 用户停止下载后隐藏进度条
                    self.progress.hide()
                else:
                    self.progress.setFormat(f"%v / %m (完成)")
                self.progress.setStyleSheet(
                    "QProgressBar { border: none; border-radius: 0px; background: #f0f0f0; text-align: center; }"
                    "QProgressBar::chunk { background-color: #4CC14C; border-radius: 0px; }"
                )
                # 下载完成时不再隐藏进度条，保持显示下载完成状态
            except Exception:
                pass

        def on_export_excel(self):
            """导出Excel表格"""
            # 检查是否有数据
            if not hasattr(self, 'all_awemes') or not self.all_awemes:
                QtWidgets.QMessageBox.warning(self, '提示', '没有作品数据可以导出')
                return
            
            # 更改按钮文字和状态
            self.export_excel_btn.setText('正在导出')
            self.export_excel_btn.setEnabled(False)
            
            all_awemes_copy = list(self.all_awemes) # 传递副本
            nickname = self.nickname_label.text() or '抖音用户'
            base_folder = cfg.get('path', '') or os.getcwd()

            # 在后台线程中执行导出操作
            export_thread = threading.Thread(
                target=self.worker.export_excel, 
                args=(all_awemes_copy, nickname, base_folder), 
                daemon=True
            )
            export_thread.start()

        def _on_export_finished(self, filepath):
            """导出完成后的UI更新"""
            self.export_excel_btn.setText('导出Excel')
            self.export_excel_btn.setEnabled(True)
            msg_box = QtWidgets.QMessageBox(self)
            msg_box.setWindowTitle('导出成功')
            msg_box.setText(f'Excel文件已保存至:\n{filepath}')
            msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok)
            ok_button = msg_box.button(QtWidgets.QMessageBox.StandardButton.Ok)
            if ok_button: ok_button.setText('确认')
            msg_box.exec()

        def _on_export_error(self, error_msg):
            """导出失败后的UI更新"""
            self.export_excel_btn.setText('导出Excel表格')
            self.export_excel_btn.setEnabled(True)
            QtWidgets.QMessageBox.warning(self, '导出失败', error_msg)
            self.append_log(error_msg)

        def on_tasks_received(self, vtasks, itasks, nickname, aweme_list):
            """接收 Worker 增量获取到的作品任务"""
            # 确保进度条隐藏
            self.progress.hide()

            # 1. 更新昵称
            self.nickname_label.setText(nickname or '')
            
            self.current_nickname = nickname or ''
            # 移除此处的 save_config 逻辑

            # 2. 累积任务
            if not hasattr(self, 'vtasks_all'): self.vtasks_all = []
            if not hasattr(self, 'itasks_all'): self.itasks_all = []
            self.vtasks_all.extend(vtasks or [])
            self.itasks_all.extend(itasks or [])
            
            # 2.1 累积原始aweme数据
            # 使用Worker中累积的所有aweme数据，而不是只使用当前批次的数据
            if hasattr(self.worker, 'all_awemes'):
                self.all_awemes = self.worker.all_awemes
            else:
                if not hasattr(self, 'all_awemes'): self.all_awemes = []
                self.all_awemes.extend(aweme_list or [])

            # 辅助函数：判断类型
            def get_type_display(desc, is_image):
                if not is_image: return '视频'
                if desc and isinstance(desc, str):
                    if re.search(r'_live\d*$', desc) or '_live' in desc:
                        return '实况'
                    if re.search(r'_p\d+$', desc):
                        return '图片'
                return '图片'

            # 3. 批量添加到列表以提高性能
            items_to_add = []
            idx = self.tree.topLevelItemCount() + 1
            
            for t in (vtasks or []):
                date_display = t.get('date', '')
                desc_display = t.get('desc', '')
                item = QtWidgets.QTreeWidgetItem([
                    ' ', str(idx), date_display, desc_display, 
                    t.get('mix_name') or '', '视频'
                ])
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                item.setCheckState(0, Qt.CheckState.Unchecked)
                item.setTextAlignment(0, int(Qt.AlignmentFlag.AlignLeft))
                item.setData(0, Qt.ItemDataRole.UserRole, (t, False)) # (task, is_image=False)
                items_to_add.append(item)
                idx += 1

            for t in (itasks or []):
                date_display = t.get('date', '')
                desc_display = t.get('desc', '')
                kind = get_type_display(desc_display, True)
                item = QtWidgets.QTreeWidgetItem([
                    ' ', str(idx), date_display, desc_display, 
                    t.get('mix_name') or '', kind
                ])
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                item.setCheckState(0, Qt.CheckState.Unchecked)
                item.setTextAlignment(0, int(Qt.AlignmentFlag.AlignLeft))
                item.setData(0, Qt.ItemDataRole.UserRole, (t, True)) # (task, is_image=True)
                items_to_add.append(item)
                idx += 1

            # 批量添加项目到树形控件
            if items_to_add:
                self.tree.setUpdatesEnabled(False)  # 暂时禁用更新以提高性能
                try:
                    self.tree.addTopLevelItems(items_to_add)
                finally:
                    self.tree.setUpdatesEnabled(True)  # 重新启用更新

            # 4. 同步旧的 vtasks/itasks 变量（保持兼容）
            self.vtasks = list(self.vtasks_all)
            self.itasks = list(self.itasks_all)
            
            # 强制刷新界面
            self.tree.repaint()

        def showEvent(self, a0):
            """窗口显示事件（用于修复列宽）"""
            try:
                header = self.tree.header()
                fm = self.tree.fontMetrics()
                target_px = fm.horizontalAdvance('汉' * 6) + 12
                if header:
                    header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeMode.Fixed)
                    header.resizeSection(4, int(target_px))
            except Exception:
                pass
            return super().showEvent(a0)

        def resizeEvent(self, a0):
            """窗口大小调整事件（用于修复列宽）"""
            try:
                header = self.tree.header()
                fm = self.tree.fontMetrics()
                target_px = fm.horizontalAdvance('汉' * 4) + 12
                if header:
                    header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeMode.Fixed)
                    header.resizeSection(4, int(target_px))
            except Exception:
                pass
            return super().resizeEvent(a0)
        
        def on_show_user_list(self):
            """显示用户列表窗口"""
            try:
                if self.user_list_window:
                    self.user_list_window.load_users() # 重新加载以防配置变更
                    self.user_list_window.show()
                    self.user_list_window.raise_()
                    self.user_list_window.activateWindow()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, '错误', f'无法打开用户列表: {e}')

        def on_fetch(self):
            """获取作品 / 停止获取"""
            # 1. 检查是否正在获取（切换为停止）
            if self.fetch_btn.text() == '停止获取':
                try:
                    if hasattr(self.worker, '_fetch_stop_requested'):
                        self.worker._fetch_stop_requested = True
                    self.append_log('[信息] 已请求停止获取')
                except Exception:
                    pass
                return
            
            # 2. 开始获取
            url = self.url_edit.text().strip()
            if not url:
                QtWidgets.QMessageBox.warning(self, '提示', '请输入主页链接')
                return
            cookie = cfg.get('cookie', '')
            if not cookie:
                QtWidgets.QMessageBox.warning(self, '提示', '请在设置中配置 Cookie')
                return
            
            # 3. 禁用除停止获取按钮外的所有按钮
            self.url_label_btn.setEnabled(False)
            self.settings_btn.setEnabled(False)
            self.clear_btn.setEnabled(False)
            self.select_all_btn.setEnabled(False)
            self.export_excel_btn.setEnabled(False)
            self.invert_btn.setEnabled(False)
            self.download_btn.setEnabled(False)
            
            # 4. 清理旧数据
            try:
                self.tree.clear()
                self.vtasks_all = []
                self.itasks_all = []
                self.vtasks = []
                self.itasks = []
                self.all_awemes = []  # 清空aweme数据
                self.current_nickname = '' 
                if hasattr(self.worker, 'all_awemes'): self.worker.all_awemes = []  # 清空Worker中的aweme数据
                if hasattr(self.worker, '_completed_tasks'): self.worker._completed_tasks = []
                if hasattr(self.worker, '_failed_tasks'): self.worker._failed_tasks = []
                if hasattr(self.worker, '_total_received'): self.worker._total_received = 0
                self.progress.setValue(0)
                self.progress.hide()
                self.status.setText('')
                self.append_log('[信息] 已清空上次获取的列表')
            except Exception:
                pass
            
            # 5. 启动后台获取
            self.fetch_btn.setText('停止获取')
            self.fetch_btn.setEnabled(True)
            self.fetch_btn.setProperty("running", True)
            self.style().unpolish(self.fetch_btn) # 刷新样式
            self.style().polish(self.fetch_btn)   # 刷新样式
            
            
            # 6. 重置停止标志并启动
            self.worker._fetch_stop_requested = False
            self._thread = threading.Thread(target=self.worker.fetch_tasks, args=(url, cookie), daemon=True)
            self._thread.start()

        def closeEvent(self, a0):
            """窗口关闭事件"""
            running_tasks = False
            if hasattr(self, '_thread') and self._thread and self._thread.is_alive():
                running_tasks = True
            elif hasattr(self.worker, '_pause_requested') and getattr(self.worker, '_pause_requested', False):
                running_tasks = True
            
            if running_tasks:
                msg_box = QtWidgets.QMessageBox(self)
                msg_box.setWindowTitle('确认退出')
                msg_box.setText('检测到有下载任务正在运行，确定要关闭程序吗？\n\n注意：关闭程序将终止所有正在进行的下载任务。')
                msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Cancel)
                msg_box.setDefaultButton(QtWidgets.QMessageBox.StandardButton.Cancel)
                ok_button = msg_box.button(QtWidgets.QMessageBox.StandardButton.Ok)
                cancel_button = msg_box.button(QtWidgets.QMessageBox.StandardButton.Cancel)
                if ok_button: ok_button.setText('确认退出')
                if cancel_button: cancel_button.setText('取消')
                
                if msg_box.exec() != QtWidgets.QMessageBox.StandardButton.Ok:
                    if a0:
                        a0.ignore()
                    return
            
            # 确认关闭
            try:
                # 请求停止所有任务
                if hasattr(self.worker, '_fetch_stop_requested'):
                    self.worker._fetch_stop_requested = True
                if hasattr(self.worker, '_pause_requested'):
                    self.worker._pause_requested = True
                if hasattr(self.worker, '_download_stop_requested'):
                    self.worker._download_stop_requested = True
                
                # 等待线程（最多5秒）
                if hasattr(self, '_thread') and self._thread and self._thread.is_alive():
                    self._thread.join(timeout=5.0)
                    
                # 关闭所有子窗口
                for w in (self.log_window, self.user_list_window, self.settings_window):
                    if w: w.close()
                    
            except Exception as e:
                print(f"[警告] 关闭时清理资源出现异常: {e}")
            
            if a0:
                a0.accept()

        def on_download(self):
            """开始下载按钮处理"""
            # 检查是否正在下载（切换为停止）
            if self.download_btn.text() == '停止下载':
                try:
                    if hasattr(self.worker, '_download_stop_requested'):
                        self.worker._download_stop_requested = True
                    self.append_log('[信息] 已请求停止下载')
                    # 立即更新按钮状态
                    self.download_btn.setText('开始下载')
                    
                    # 移除 setStyleSheet
                    # self.download_btn.setStyleSheet('''...''')
                    # 设置 "running" 属性为 False，QSS会自动应用蓝色样式
                    self.download_btn.setProperty("running", False)
                    self.style().unpolish(self.download_btn) # 刷新样式
                    self.style().polish(self.download_btn)   # 刷新样式
                    
                    # 点击停止下载后隐藏进度条
                    self.progress.hide()
                except Exception:
                    pass
                return
            
            # 1. 获取选中的任务
            selected = []
            for i in range(self.tree.topLevelItemCount()):
                it = self.tree.topLevelItem(i)
                if it and it.checkState(0) == Qt.CheckState.Checked:
                    data = it.data(0, Qt.ItemDataRole.UserRole)
                    if data:
                        selected.append(data)
                        
            if not selected:
                QtWidgets.QMessageBox.warning(self, '提示', '请先选择要下载的作品')
                return
            
            sel_v = [d[0] for d in selected if not d[1]] # (task, is_image=False)
            sel_i = [d[0] for d in selected if d[1]] # (task, is_image=True)
            
            # 2. 确定保存路径
            base_folder = cfg.get('path', '') or os.getcwd()
            nickname_for_folder = self.nickname_label.text() or 'Douyin_User'  # 移除了对last_nickname字段的依赖
            user_folder = os.path.join(base_folder, sanitize_filename(nickname_for_folder) or 'Douyin_Downloads')
            
            if not safe_mkdir(user_folder):
                QtWidgets.QMessageBox.critical(self, '错误', f'创建目录失败: {user_folder}')
                return
                
            # 3. 获取配置
            threads = int(cfg.get('threads', DEFAULT_THREAD_COUNT))
            use_mix_folder = cfg.get('use_mix_folder', True)
            include_date = cfg.get('include_date_in_filename', True)

            # 4. 预处理任务（应用设置）
            def apply_settings_to_tasks(tasks, is_image):
                out = []
                for t in tasks:
                    nt = dict(t) # 创建副本
                    # 不使用合集
                    if not use_mix_folder:
                        nt['mix_name'] = None
                    
                    # 不再处理 desc 字符串，而是将配置存入 task
                    nt['include_date_in_filename'] = include_date
                    
                    out.append(nt)
                return out

            sel_v_proc = apply_settings_to_tasks(sel_v, False)
            sel_i_proc = apply_settings_to_tasks(sel_i, True)
            
            # 5. 启动后台下载
            # 显示并重置进度条
            self.progress.show()
            self.progress.setMaximum(max(1, len(sel_v_proc) + len(sel_i_proc)))
            self.progress.setValue(0)
            self.on_progress(0, max(1, len(sel_v_proc) + len(sel_i_proc))) # 恢复蓝色
            
            # 重置停止标志
            self.worker._download_stop_requested = False
            self.worker._pause_requested = False

            # 禁用所有按钮
            self.url_label_btn.setEnabled(False)
            self.settings_btn.setEnabled(False)
            self.clear_btn.setEnabled(False)
            self.select_all_btn.setEnabled(False)
            self.export_excel_btn.setEnabled(False)
            self.invert_btn.setEnabled(False)
            self.fetch_btn.setEnabled(False)
            
            # 设置下载按钮为停止下载按钮
            self.download_btn.setText('停止下载')
            self.download_btn.setEnabled(True)
            self.download_btn.setProperty("running", True)
            self.style().unpolish(self.download_btn) # 刷新样式
            self.style().polish(self.download_btn)   # 刷新样式

            self._thread = threading.Thread(
                target=self.worker.download_tasks, 
                args=(sel_v_proc, sel_i_proc, user_folder, threads), 
                daemon=True
            )
            self._thread.start()

        def on_fetch_finished(self):
            """获取完成处理（用于自动全选 和 保存用户）"""
            try:
                url = self.url_edit.text().strip()
                nickname = self.current_nickname # 使用暂存的昵称
                
                if url and nickname:
                    # 提取当前URL的sec_user_id
                    current_sec_user_id = extract_sec_user_id_from_url(url)
                    if current_sec_user_id:
                        users = cfg.get('users', [])
                        existing_user = None
                        existing_user_index = -1
                        
                        # 通过sec_user_id匹配用户
                        for idx, user in enumerate(users):
                            user_url = user.get('url', '')
                            user_sec_user_id = extract_sec_user_id_from_url(user_url)
                            if user_sec_user_id == current_sec_user_id:
                                existing_user = user
                                existing_user_index = idx
                                break
                        
                        if not existing_user:
                            # 添加新用户
                            users.append({'username': nickname, 'url': url})
                            cfg['users'] = users
                            save_config(cfg)
                            self.append_log(f'[信息] 已保存用户: {nickname}')
                        elif not existing_user.get('username') or existing_user.get('username') != nickname:
                            # 更新用户名（如果为空或不匹配）
                            users[existing_user_index]['username'] = nickname
                            cfg['users'] = users
                            save_config(cfg)
                            self.append_log(f'[信息] 已更新用户: {nickname}')
            except Exception as e:
                self.append_log(f'[警告] 保存用户信息失败: {e}')
            
            # 自动全选逻辑
            try:
                if bool(cfg.get('auto_select_after_fetch', False)):
                    # 自动全选
                    self.on_select_all()
                    self.append_log('[信息] 获取完成，已自动全选')
                else:
                    self.append_log('[信息] 获取完成')
            except Exception as e:
                self.append_log(f'[警告] 获取完成处理失败: {e}')

        def on_settings(self):
            """显示设置窗口"""
            try:
                if self.settings_window:
                    self.settings_window.show()
                    self.settings_window.raise_()
                    self.settings_window.activateWindow()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, '错误', f'无法打开设置窗口: {e}')
        
        def show_first_time_settings(self):
            """首次启动时显示设置窗口"""
            msg_box = QtWidgets.QMessageBox(self)
            msg_box.setWindowTitle('欢迎使用')
            msg_box.setText(
                '欢迎使用抖音主页作品批量下载工具！\n\n'
                '检测到这是您第一次使用本程序，\n'
                '请先配置 Cookie 和保存路径。\n\n'
                '点击“确认”将打开设置窗口。'
            )
            msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok)
            ok_button = msg_box.button(QtWidgets.QMessageBox.StandardButton.Ok)
            if ok_button: ok_button.setText('确认')
            msg_box.exec()
            
            self.on_settings()
        
        def on_status_click(self, event):
            """点击状态标签打开日志窗口"""
            try:
                if self.log_window:
                    self.log_window.show()
                    self.log_window.raise_()
                    self.log_window.activateWindow()
            except Exception:
                pass

        def on_select_all(self):
            """全选"""
            self._programmatic_change = True
            try:
                for i in range(self.tree.topLevelItemCount()):
                    it = self.tree.topLevelItem(i)
                    if it:
                        it.setCheckState(0, Qt.CheckState.Checked)
                        it.setSelected(True)
            finally:
                self._programmatic_change = False
            self.update_status_label()

        def on_invert(self):
            """反选"""
            self._programmatic_change = True
            try:
                for i in range(self.tree.topLevelItemCount()):
                    it = self.tree.topLevelItem(i)
                    if it:
                        current_state = it.checkState(0)
                        new_state = Qt.CheckState.Unchecked if current_state == Qt.CheckState.Checked else Qt.CheckState.Checked
                        it.setCheckState(0, new_state)
                        it.setSelected(new_state == Qt.CheckState.Checked)
            finally:
                self._programmatic_change = False
            self.update_status_label()

        def on_clear_list(self):
            """清空列表"""
            msg_box = QtWidgets.QMessageBox(self)
            msg_box.setWindowTitle('确认')
            msg_box.setText('确定要清空当前列表吗？此操作不会删除已下载的文件。')
            msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Cancel)
            msg_box.setDefaultButton(QtWidgets.QMessageBox.StandardButton.Cancel)
            ok_button = msg_box.button(QtWidgets.QMessageBox.StandardButton.Ok)
            cancel_button = msg_box.button(QtWidgets.QMessageBox.StandardButton.Cancel)
            if ok_button: ok_button.setText('确认')
            if cancel_button: cancel_button.setText('取消')
            
            if msg_box.exec() != QtWidgets.QMessageBox.StandardButton.Ok:
                return
            
            try:
                self.tree.clear()
                self.vtasks_all = []
                self.itasks_all = []
                self.vtasks = []
                self.itasks = []
                self.all_awemes = []  # 同时清空aweme数据
                self.current_nickname = ''
                self.progress.setValue(0)
                self.progress.hide()
            except Exception:
                pass
            
            self.append_log('[信息] 已清空当前列表')

        def on_worker_finished(self):
            """工作线程完成处理（Fetch 或 Download）"""
            # 1. 恢复所有按钮
            self.url_label_btn.setEnabled(True)
            self.settings_btn.setEnabled(True)
            self.clear_btn.setEnabled(True)
            self.select_all_btn.setEnabled(True)
            if OPENPYXL_AVAILABLE:
                self.export_excel_btn.setEnabled(True)
                
            self.invert_btn.setEnabled(True)
            self.download_btn.setEnabled(True)
            self.fetch_btn.setEnabled(True)
            self.fetch_btn.setText('获取作品')
            self.download_btn.setText('开始下载')
            
            # 设置 "running" 属性为 False，QSS会自动应用蓝色样式
            self.fetch_btn.setProperty("running", False)
            self.style().unpolish(self.fetch_btn) # 刷新样式
            self.style().polish(self.fetch_btn)   # 刷新样式
            
            self.download_btn.setProperty("running", False)
            self.style().unpolish(self.download_btn) # 刷新样式
            self.style().polish(self.download_btn)   # 刷新样式
            
            
            # 4. 不再隐藏进度条，保持显示下载完成状态
            # self.progress.hide()
            
            # 5. 如果进度条是满的，确保是绿色
            if self.progress.value() == self.progress.maximum() and self.progress.maximum() > 0:
                self.on_download_finished()
            
            # 如果是用户主动停止下载，也调用下载完成的处理
            if hasattr(self.worker, '_download_stop_requested') and self.worker._download_stop_requested:
                self.on_download_finished()
            # 如果是用户主动停止下载，隐藏进度条
            elif hasattr(self.worker, '_download_stop_requested') and self.worker._download_stop_requested:
                self.progress.hide()


    # ========================================
    # GUI 启动入口
    # ========================================
    app = QtWidgets.QApplication(sys.argv)
    
    try:
        with tempfile.NamedTemporaryFile(suffix='.ico', delete=False) as tmp:
            tmp.write(ICON_BYTES)
            tmp_icon_path = tmp.name
        app_icon = QtGui.QIcon(tmp_icon_path)
        app.setWindowIcon(app_icon)
    except Exception as e:
        print(f"Warning: Failed to create temp icon: {e}")

    checkmark_path = ''
    try:
        checkmark_svg = b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 16 16"><path fill="white" stroke="white" stroke-width="0.5" d="M13.5 4l-7 7-3.5-3.5 1.5-1.5 2 2 5.5-5.5z"/></svg>'
        with tempfile.NamedTemporaryFile(suffix='.svg', delete=False, mode='wb') as tmp_check:
            tmp_check.write(checkmark_svg)
            checkmark_path = tmp_check.name.replace('\\', '/')
    except Exception as e:
        print(f"Warning: Failed to create temp checkmark svg: {e}")

    # 全局样式表
    try:
        app.setStyleSheet("""
        /* ---------------- 按钮 ---------------- */
        QPushButton {
            background-color: #409EFF; border: 1px solid #409EFF; color: white;
            padding: 6px 14px; border-radius: 0px; font-weight: 500; font-size: 13px;
        }
        QPushButton:hover { background-color: #66b1ff; border: 1px solid #66b1ff; }
        QPushButton:pressed { background-color: #3a8ee6; border: 1px solid #3a8ee6; }
        QPushButton:disabled {
            background-color: #a0cfff; border: 1px solid #a0cfff; color: #f0f0f0;
        }

        /* 添加 [running="true"] 状态，用于“停止”按钮 */
        /* 这个样式与 clear_btn 的红色样式一致 */
        QPushButton[running="true"] {
            background: #d9534f; color: white; padding: 7px 14px;
            border: 1px solid #d9534f; font-weight: 500; font-size: 13px;
        }
        QPushButton[running="true"]:hover { 
            background: #fa8480; border: 1px solid #fa8480;
        }
        QPushButton[running="true"]:pressed { 
            background: #d9534f; border: 1px solid #d9534f;
        }
        QPushButton[running="true"]:disabled { 
            background: #f0b3b3; color: #f8e6e6; border: 1px solid #f0b3b3;
        }


        /* ---------------- 输入框 ---------------- */
        QLineEdit, QTextEdit, QSpinBox {
            border: 1px solid #dcdfe6; background: #ffffff; color: #303133;
            padding: 6px; border-radius: 0px; font-size: 13px;
        }
        QLineEdit:focus, QTextEdit:focus, QSpinBox:focus {
            border: 1px solid #409EFF; background: #f9fcff;
        }

        /* ---------------- 复选框 (全局) ---------------- */
        QTreeView::indicator, QTreeWidget::indicator {
            width: 16px; height: 16px; border: 1px solid #c0c4cc;
            border-radius: 2px; background: #ffffff;
        }
        QTreeView::indicator:hover, QTreeWidget::indicator:hover {
            border: 1px solid #409EFF;
        }
        QTreeView::indicator:checked, QTreeWidget::indicator:checked {
            background-color: #409EFF; border: 1px solid #409EFF;
            image: url(""" + checkmark_path + r""");
        }

        /* ---------------- 列表 QTreeWidget ---------------- */
        QTreeWidget {
            background: #ffffff; border: 1px solid #e4e7ed;
            alternate-background-color: #fafbfc; gridline-color: #f2f6fc;
            selection-background-color: #d9eaff; font-size: 13px;
        }
        QTreeWidget::item { padding: 6px 4px; color: #222222; }
        QTreeWidget::item:hover { background: #f3f8fe; }
        QTreeWidget::item:selected { background: #cfe4ff; color: #000000; }

        /* ---------------- 进度条 ---------------- */
        QProgressBar {
            border: 1px solid #dcdfe6; background: #f5f7fa; height: 22px;
            border-radius: 0px; text-align: center; font-size: 12px; color: #303133;
        }
        QProgressBar::chunk {
            background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #66b1ff, stop:1 #409EFF);
            border-radius: 0px;
        }

        /* ---------------- 滚动条 ---------------- */
        QScrollBar:vertical {
            border: none; background: #f5f7fa; width: 10px; margin: 0px;
        }
        QScrollBar::handle:vertical {
            background: #c0c4cc; border-radius: 0px; min-height: 20px;
        }
        QScrollBar::handle:vertical:hover { background: #a6a9ad; }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            height: 0px; background: none;
        }
        QScrollBar:horizontal {
            border: none; background: #f5f7fa; height: 10px; margin: 0px;
        }
        QScrollBar::handle:horizontal {
            background: #c0c4cc; border-radius: 0px; min-width: 20px;
        }
        QScrollBar::handle:horizontal:hover { background: #a6a9ad; }
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
            width: 0px; background: none;
        }
        """)
    except Exception as e:
        print(f"Warning: Failed to set stylesheet: {e}")

    w = MainWindow(checkmark_path)
    w.show()
    app.exec()


if __name__ == "__main__":
    run_gui()